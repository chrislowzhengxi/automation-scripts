import openpyxl
from openpyxl.utils.cell import column_index_from_string
from openpyxl.utils import get_column_letter
from openpyxl.styles import Border, Side
from copy import copy 

import os
import sys



# Setup logging
# log_path = os.path.join(os.environ['USERPROFILE'], "Downloads\\Automation\\revenue_update", "log.txt")
log_path = os.path.join(os.path.dirname(__file__), "log.txt")
log = open(log_path, "w", encoding="utf-8")

def logprint(msg):
    print(msg)
    log.write(msg + "\n")


# Get year_month from Power Automate argument
if len(sys.argv) < 2:
    logprint("Error: YearMonth argument missing.")
    sys.exit(1)

year_month = sys.argv[1]
logprint(f"Year-Month received: {year_month}")

# Get Taiwan year and month
taiwan_year = str(int(year_month[:4]) - 1911)  # e.g., 2025 -> 114
month = year_month[4:]                         # e.g., "04"
taiwan_year_month = taiwan_year + month        # e.g., "11404"

downloads_folder = os.path.join(os.environ['USERPROFILE'], "Downloads")
rptis_path = os.path.join(downloads_folder, f"RPTIS10_I_A01_{year_month}.xlsx")
monthly_path = os.path.join(downloads_folder, f"High-Monthy-營收公告-{taiwan_year_month}.xlsx")
if not os.path.exists(monthly_path):
    logprint(f"Expected file not found: {monthly_path}")
    # Try to find the most recent matching "High-Monthy" file
    candidates = [
        f for f in os.listdir(downloads_folder)
        if f.startswith("High-Monthy-營收公告-") and f.endswith(".xlsx")
    ]
    if candidates:
        # Pick the most recently modified file and copy it as the expected name
        candidates.sort(key=lambda f: os.path.getmtime(os.path.join(downloads_folder, f)), reverse=True)
        fallback_file = os.path.join(downloads_folder, candidates[0])
        logprint(f"Fallback to latest file: {fallback_file}")
        
        # Copy fallback file to new expected filename
        from shutil import copyfile
        copyfile(fallback_file, monthly_path)
        logprint(f"Copied fallback to expected name: {monthly_path}")
    else:
        logprint("No fallback file found.")
        sys.exit(1)


logprint(f"RPTIS file path: {rptis_path}")
logprint(f"Monthly tracker path: {monthly_path}")

# Check if files exist
if not os.path.exists(rptis_path):
    logprint(f"Error: RPTIS file not found at {rptis_path}")
    sys.exit(1)

if not os.path.exists(monthly_path):
    logprint(f"Error: Monthly tracker file not found at {monthly_path}")
    sys.exit(1)


# Load source workbook (report)
rptis_wb = openpyxl.load_workbook(rptis_path, data_only=True)
rptis_ws = rptis_wb.active
new_revenue = rptis_ws["B9"].value

# Convert to int if string (e.g. "3,575,188")
if isinstance(new_revenue, str):
    new_revenue = int(new_revenue.replace(",", "").strip())
new_revenue = round(new_revenue / 1000)


# Load monthly tracker workbook
monthly_wb = openpyxl.load_workbook(monthly_path)
monthly_ws = monthly_wb.active

# Row setup
row_cumulative = 25  # cumulative revenue
row_monthly = 23     # monthly difference

# Find next empty column in row 25 (start at column C = index 3)
col = 3
while monthly_ws.cell(row=row_cumulative, column=col).value:
    col += 1

# Get column letters
col_letter = get_column_letter(col)
prev_col_letter = get_column_letter(col - 1)


# Write new column: 
# Write new revenue to row 25
cell = monthly_ws[f"{col_letter}{row_cumulative}"] 
cell.value = new_revenue
cell.number_format = '#,##0'  # Format as number with commas

# Read previous month's revenue
prev_value = monthly_ws[f"{prev_col_letter}{row_cumulative}"].value
if isinstance(prev_value, str):
    prev_value = int(prev_value.replace(",", "").strip())

# Compute difference
net_revenue = new_revenue - prev_value

# Write net difference to row 23
diff_cell = monthly_ws[f"{col_letter}{row_monthly}"] 
diff_cell.value = net_revenue
diff_cell.number_format = '#,##0'  # Format as number with commas



# Write taiwan year-month to row 2
monthly_ws[f"{col_letter}2"] = int(taiwan_year_month)
# Assign formulas to row 24 and 26 based on 12 columns earlier than current
col_12_back = get_column_letter(col - 12)

monthly_ws[f"{col_letter}24"] = f"={col_12_back}23"
monthly_ws[f"{col_letter}26"] = f"={col_12_back}25"

monthly_ws[f"{col_letter}24"].number_format = '#,##0'
monthly_ws[f"{col_letter}26"].number_format = '#,##0'


# Copy remaining rows from 27 to 134
for row in range(27, 135):
    source = monthly_ws.cell(row=row, column=col - 1)
    target = monthly_ws.cell(row=row, column=col)
    target.value = source.value
    target.number_format = source.number_format

def copy_cell(source, target, copy_value=False):
    """
    Copies formatting and value (if applicable)
    """
    if copy_value:
        target.value = source.value
    target.number_format = source.number_format
    target.font = copy(source.font)
    target.fill = copy(source.fill)
    target.border = copy(source.border)
    target.alignment = copy(source.alignment)

for row in [2, 23, 24, 25, 26]:
    source = monthly_ws.cell(row=row, column=col - 1)
    target = monthly_ws.cell(row=row, column=col)
    copy_cell(source, target)

for row in range(27, 135): 
    source = monthly_ws.cell(row=row, column=col - 1)
    target = monthly_ws.cell(row=row, column=col)
    copy_cell(source, target, copy_value=True)

thin = Side(border_style="thin", color="000000")
bold = Side(border_style="medium", color="000000")
# 1) Reset the old “newest” column (prev_col_letter) back to thin all around:
for row in range(2, 135):
    old_cell = monthly_ws[f"{prev_col_letter}{row}"]
    old_cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)

# 2) Now apply the bold right-edge only to your actual newest column (col_letter):
for row in range(2, 135):
    new_cell = monthly_ws[f"{col_letter}{row}"]
    new_cell.border = Border(left=thin, right=bold, top=thin, bottom=thin)


# Define the actual first data column (e.g., "EI" = column 139)
DATA_START_COL = column_index_from_string("EI")  # 139
KEEP_START = col - 12                            # e.g. for 11404 col=139, KEEP_START=127 (11304)

# 1) Hide all columns older than KEEP_START
for c in range(DATA_START_COL, KEEP_START):
    letter = get_column_letter(c)
    monthly_ws.column_dimensions[letter].hidden = True

# 2) Ensure your 13 most‐recent are visible (optional but safe)
for c in range(KEEP_START, col + 1):
    letter = get_column_letter(c)
    monthly_ws.column_dimensions[letter].hidden = False



# Save the file
monthly_wb.save(monthly_path)

logprint(f"✅ Success. Net revenue for {year_month}: {net_revenue:,}")
log.close()