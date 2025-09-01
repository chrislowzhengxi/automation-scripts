#!/usr/bin/env python3
from __future__ import annotations
import argparse
from pathlib import Path
from copy import copy as copy_style
from datetime import datetime, date
from openpyxl.styles import Alignment
from openpyxl.styles import PatternFill
import re
import sys
import pandas as pd
import openpyxl
from datetime import datetime, date
from openpyxl.utils import get_column_letter

YELLOW_FILL = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
# ---------------- helpers ----------------

def norm_code(x):
    if pd.isna(x):
        return None
    s = str(x).strip().replace(",", "")
    if s.endswith(".0"):
        s = s[:-2]
    return s or None

def sanitize_sheet_name(name: str) -> str:
    # Excel sheet name rules: <=31 chars, cannot contain : \ / ? * [ ]
    name = re.sub(r'[:\\/\?\*\[\]]', ' ', name)
    return name[:31] if len(name) > 31 else name

def choose_output_path(export_path: Path, inplace: bool, output: str | None) -> Path:
    if inplace:
        return export_path
    if output:
        return Path(output)
    return export_path.with_name(export_path.stem + "_grouped.xlsx")

def read_export_frame(export_path: Path, sheet_name: str | None) -> tuple[pd.DataFrame, str]:
    if sheet_name is None:
        xls = pd.ExcelFile(export_path)
        sheet_name = xls.sheet_names[0]
    df = pd.read_excel(export_path, sheet_name=sheet_name, dtype=object)
    df.columns = [str(c).strip() for c in df.columns]
    return df, sheet_name

def find_gl_column(df: pd.DataFrame) -> str:
    for c in df.columns:
        if c.strip() == "G/L科目":
            return c
    raise ValueError("Could not find column 'G/L科目' in the export.")

def load_mapping(mapping_path: Path) -> dict[str, str]:
    df_map_raw = pd.read_excel(mapping_path, sheet_name=0, dtype=object, usecols=[0, 1], header=0)
    df_map_raw = df_map_raw.rename(columns={df_map_raw.columns[0]: "number", df_map_raw.columns[1]: "name"})
    df_map_raw["number_norm"] = df_map_raw["number"].apply(norm_code)
    df_map = df_map_raw.dropna(subset=["number_norm"]).copy()
    return dict(zip(df_map["number_norm"], df_map["name"].fillna("").astype(str)))

def pick_columns_B_to_X(df: pd.DataFrame) -> list[str]:
    bx_end = min(24, len(df.columns))  # B..X inclusive
    return [c for c in df.columns[1:bx_end] if c != "_code"]

def ensure_unique_title(base_title: str, taken: set[str]) -> str:
    title = sanitize_sheet_name(base_title) or "Sheet"
    if title not in taken:
        return title
    suffix = 1
    while True:
        candidate = sanitize_sheet_name(f"{title[:25]}_{suffix}")
        if candidate not in taken:
            return candidate
        suffix += 1

def to_date_value(v):
    """Return a python date (preferred) for Excel, or original value if not parseable."""
    if isinstance(v, (datetime, date)):
        return v.date() if isinstance(v, datetime) else v
    # try common string/number inputs
    try:
        dt = pd.to_datetime(v, errors="coerce")
        if pd.isna(dt):
            return v
        # pandas returns Timestamp -> convert to date
        return dt.date()
    except Exception:
        return v


def copy_header_style(src_ws, src_col_indexes: list[int], dst_ws, dst_row: int = 1):
    """
    Copy header cell style from src_ws row=1 columns (by numeric index),
    and apply to dst_ws row=dst_row 1..N.
    """
    for j, src_col_idx in enumerate(src_col_indexes, start=1):
        src_cell = src_ws.cell(row=1, column=src_col_idx)
        dst_cell = dst_ws.cell(row=dst_row, column=j)   # <-- use dst_row here
        if src_cell.has_style:
            dst_cell.font = copy_style(src_cell.font)
            dst_cell.fill = copy_style(src_cell.fill)
            dst_cell.border = copy_style(src_cell.border)
            dst_cell.alignment = copy_style(src_cell.alignment)
            dst_cell.number_format = src_cell.number_format
            dst_cell.protection = copy_style(src_cell.protection)
        # column width copy stays the same (width is per-column)
        try:
            src_letter = openpyxl.utils.get_column_letter(src_col_idx)
            dst_letter = openpyxl.utils.get_column_letter(j)
            width = src_ws.column_dimensions[src_letter].width
            if width:
                dst_ws.column_dimensions[dst_letter].width = width
        except Exception:
            pass

def _safe_col(ws, idx: int) -> str | None:
    """Return Excel letter for idx if it exists on the sheet, else None."""
    if idx <= 0 or idx > ws.max_column:
        return None
    return get_column_letter(idx)

def _group_cols(ws, start_idx: int, end_idx: int, hidden=True, outline_level=1):
    """Group columns if they exist (inclusive)."""
    start_letter = _safe_col(ws, start_idx)
    end_letter   = _safe_col(ws, end_idx)
    if start_letter and end_letter:
        ws.column_dimensions.group(start=start_letter, end=end_letter,
                                   outline_level=outline_level, hidden=hidden)
        # make sure outline symbols are visible
        ws.sheet_properties.outlinePr.summaryBelow = True
        ws.sheet_view.showOutlineSymbols = True

def _apply_common_sheet_format(ws):
    """Freeze top row + enable filter on the whole used range."""
    # Freeze first row
    ws.freeze_panes = "A2"
    # AutoFilter over the full used range
    last_col = get_column_letter(ws.max_column or 1)
    last_row = ws.max_row or 1
    ws.auto_filter.ref = f"A1:{last_col}{last_row}"

def _format_column_N(ws):
    """Number format for column N: #,##0;[Red](#,##0) on data rows (row ≥ 2)."""
    if ws.max_column < 14:
        return
    fmt = '#,##0;[Red](#,##0)'
    for r in range(2, ws.max_row + 1):
        ws.cell(row=r, column=14).number_format = fmt

def _apply_groupings(ws):
    """
    Group columns with outline (+/−):
      A–B, D–E, G, K–L, U–W (best-effort if sheet has fewer columns).
    """
    # A–B
    _group_cols(ws, 1, 2, hidden=True)
    # D–E
    _group_cols(ws, 4, 5, hidden=True)
    # G (single column)
    _group_cols(ws, 7, 7, hidden=True)
    # K–L
    _group_cols(ws, 11, 12, hidden=True)
    # U–W (21–23). Will silently skip if not present.
    _group_cols(ws, 21, 23, hidden=True)
# ---------------- core ----------------

def group_export_by_account(
    export_path: Path,
    mapping_path: Path,
    output_path: Path,
    sheet_name: str | None,
    inplace: bool,
    drop_original_titles: list[str],
    date_columns: list[str],
    cutoff_date: date | None = None, 
) -> dict:
    # 1) Read export + detect columns
    df_export, sheet_used = read_export_frame(export_path, sheet_name)
    gl_col = find_gl_column(df_export)

    # 2) mapping
    number_to_name = load_mapping(mapping_path)

    # 3) normalize and filter
    df_export["_code"] = df_export[gl_col].apply(norm_code)
    df_export_valid = df_export[df_export["_code"].notna()].copy()
    df_export_valid = df_export_valid[df_export_valid["_code"].isin(number_to_name.keys())]

    # 4) columns B..X
    selected_cols = pick_columns_B_to_X(df_export)

    # we also need indexes of B..X in the source sheet to copy header style later
    src_col_indexes = [df_export.columns.get_loc(col) + 1 for col in selected_cols]  # 1-based for openpyxl

    # 5) open workbook
    wb = openpyxl.load_workbook(export_path)
    src_ws = wb[sheet_used]  # header style source
    # ---- Highlighting helpers (need selected_cols + wb) ----
    code_to_title: dict[str, str] = {}

    PREF_KEY_COLS = ["文件號碼", "過帳日期", "G/L科目", "國貨幣計算之金額"]

    def _is_blank(v) -> bool:
        # Treat None, empty, whitespace, and pandas NaN/NaT as blank
        if v is None:
            return True
        # pandas/Excel NaN -> float('nan') or string 'nan'
        try:
            import math
            if isinstance(v, float) and math.isnan(v):
                return True
        except Exception:
            pass
        s = str(v).strip()
        return s == "" or s.lower() in {"nan", "nat"}


    def _norm_scalar(x):
        if x is None:
            return ""
        try:
            dt = pd.to_datetime(x, errors="coerce")
            if not pd.isna(dt):
                return dt.date().isoformat()
        except Exception:
            pass
        s = str(x).strip()
        if s.endswith(".0"):
            s = s[:-2]
        return s

    def _make_key_from_series(sr: pd.Series, key_cols: list[str]) -> tuple:
        return tuple(_norm_scalar(sr.get(k)) for k in key_cols)

    def _make_key_from_ws_row(ws_g, r: int, idx_map: dict[str, int]) -> tuple:
        return tuple(_norm_scalar(ws_g.cell(row=r, column=cidx).value) for cidx in idx_map.values())

    def _effective_key_cols(headers: list[str]) -> list[str]:
        has = set(headers)
        if {"文件號碼", "過帳日期"} <= has:
            return ["文件號碼", "過帳日期"]
        if "文件號碼" in has:
            return ["文件號碼"]
        return [h for h in PREF_KEY_COLS if h in has][:1]

    def _col_indexes_map(headers: list[str], wants: list[str]) -> dict[str, int]:
        return {h: headers.index(h) + 1 for h in wants if h in headers}

    def _highlight_code_rows(code: str, df_sub: pd.DataFrame):
        title = code_to_title.get(code)
        if not title or title not in wb.sheetnames:
            return
        ws_g = wb[title]

        key_cols = _effective_key_cols(selected_cols)
        if not key_cols:
            return
        idx_map = _col_indexes_map(selected_cols, key_cols)

        keys = { _make_key_from_series(sr, key_cols) for _, sr in df_sub.iterrows() }

        col_cleared = selected_cols.index("結清文件") + 1 if "結清文件" in selected_cols else None

        for r in range(2, ws_g.max_row + 1):
            if col_cleared:
                cleared_val = ws_g.cell(row=r, column=col_cleared).value
                if not _is_blank(cleared_val):
                    continue
            row_key = _make_key_from_ws_row(ws_g, r, idx_map)
            if row_key in keys:
                for c in range(1, ws_g.max_column + 1):
                    ws_g.cell(row=r, column=c).fill = YELLOW_FILL

    # 6) create grouped sheets and write data
    used_titles = {ws.title for ws in wb.worksheets}
    date_cols_set = set(date_columns)

    for code, grp in df_export_valid.groupby("_code"):
        name = number_to_name.get(code, "").strip()
        base_title = f"{code} {name}".strip()
        title = ensure_unique_title(base_title, used_titles)
        used_titles.add(title)
        code_to_title[code] = title

        ws = wb.create_sheet(title=title)

        # Header values
        for j, col_name in enumerate(selected_cols, start=1):
            ws.cell(row=1, column=j, value=str(col_name))

        # Header style from export (B..X)
        copy_header_style(src_ws, src_col_indexes, ws)

        # Body rows
        for i, row in enumerate(grp[selected_cols].itertuples(index=False, name=None), start=2):
            for j, val in enumerate(row, start=1):
                header = selected_cols[j-1]
                if header in date_cols_set:
                    v = to_date_value(val)
                    ws.cell(row=i, column=j, value=v)
                    # Apply Excel date format so it shows as m/d/yyyy (no leading zero)
                    ws.cell(row=i, column=j).number_format = "m/d/yyyy"
                else:
                    ws.cell(row=i, column=j, value=val)
        
        _apply_common_sheet_format(ws)
        _apply_groupings(ws)
        _format_column_N(ws)

    # ---- Build the “>30 days 未報銷 (未結清)” summary for 預付費用 ----
    TARGET_CODES = {"12580100", "12680100"}    # 預付費用 + 其他預付款
    DATE_COL = "過帳日期"
    UNCLEARED_COL = "結清文件"                  # treat empty as 未結清

    # Ensure needed columns exist
    missing = [c for c in [DATE_COL, UNCLEARED_COL, gl_col] if c not in df_export.columns]
    if not missing:
        df_tmp = df_export.copy()

        # normalize posting date to pure date
        dt = pd.to_datetime(df_tmp[DATE_COL], errors="coerce")
        df_tmp["_posting_date"] = dt.dt.date

        co = cutoff_date or date.today()

        # 未結清 if blank/NaN
        uncleared_mask = df_tmp[UNCLEARED_COL].isna() | (df_tmp[UNCLEARED_COL].astype(str).str.strip() == "")

        # 科目 in targets (normalized)
        df_tmp["_code"] = df_tmp[gl_col].apply(norm_code)
        code_mask = df_tmp["_code"].isin(TARGET_CODES)

        # older than 30 days
        age_mask = df_tmp["_posting_date"].notna() & (
            (pd.to_datetime(co) - pd.to_datetime(df_tmp["_posting_date"]))
            .dt.days > 30
        )

        summary = df_tmp[code_mask & uncleared_mask & age_mask].copy()

        title = "說明"
        if title in wb.sheetnames:
            del wb[title]
        ws = wb.create_sheet(title=title, index=0)

        # === NEW: Section 1 header text (like the 0623 file) ===
        ws.cell(row=1, column=1, value="1. 預付費用超過30天未報銷之項目，請說明原因。")
        row_ptr = 3  # leave one blank line after the header

        # Small lead-in line (same wording as screenshot)
        ws.cell(row=row_ptr, column=1, value="→超過30天預付費用明細：")
        row_ptr += 1

        # Header for the first table starts here
        hdr_row = row_ptr
        for j, col_name in enumerate(selected_cols, start=1):
            ws.cell(row=hdr_row, column=j, value=str(col_name))
        copy_header_style(src_ws, [df_export.columns.get_loc(c) + 1 for c in selected_cols], ws, dst_row=hdr_row)

        # Rows
        data_start = hdr_row + 1
        for i, row in enumerate(summary[selected_cols].itertuples(index=False, name=None), start=data_start):
            for j, val in enumerate(row, start=1):
                ws.cell(row=i, column=j, value=val)

        # Apply m/d/yyyy to date columns if present
        DATE_COLS = {"文件日期", "過帳日期"}
        header_to_idx = {h: idx+1 for idx, h in enumerate(selected_cols)}
        for h in DATE_COLS:
            if h in header_to_idx:
                cidx = header_to_idx[h]
                for r in range(data_start, data_start + len(summary)):
                    ws.cell(row=r, column=cidx).number_format = "m/d/yyyy"
        
        # Highlight matching rows in their original grouped sheets
        # Group by GL code because the 'summary' may contain multiple codes from TARGET_CODES
        for ccode, subgrp in summary.groupby("_code"):
             _highlight_code_rows(ccode, subgrp)


                    
        # === NEW: Section 2 header text before 存出保證金 ===
        row_ptr = ws.max_row + 2  # two blank lines after the above table
        ws.cell(row=row_ptr, column=1, value="2. 存出保證金是否應取回，以及流動性分類是否正確。")
        row_ptr += 2  # one blank line before the next header

        # ---- Append 存出保證金 (split into two sections: 11780300 then 18200100) ----
        CODES_SEQ = [("11780300", "存出保證金-流動"), ("18200100", "存出保證金")]

        for idx, (code, _label) in enumerate(CODES_SEQ):
            sub = df_tmp[df_tmp["_code"] == code].copy()
            if sub.empty:
                continue

            # leave 2 blank rows after whatever content already exists
            start_row = ws.max_row + 2

            # Header (repeat for each section)
            for j, col_name in enumerate(selected_cols, start=1):
                ws.cell(row=start_row, column=j, value=str(col_name))
            copy_header_style(src_ws, [df_export.columns.get_loc(c) + 1 for c in selected_cols], ws, dst_row=start_row)

            # Rows
            for i, row in enumerate(sub[selected_cols].itertuples(index=False, name=None), start=start_row + 1):
                for j, val in enumerate(row, start=1):
                    ws.cell(row=i, column=j, value=val)

            # Date formatting for this section
            for h in DATE_COLS:
                if h in header_to_idx:
                    cidx = header_to_idx[h]
                    for r in range(start_row + 1, start_row + 1 + len(sub)):
                        ws.cell(row=r, column=cidx).number_format = "m/d/yyyy"

            # Highlight for this 保證金 code
            _highlight_code_rows(code, sub)



    # ==== Step 5 (append to 說明): >90天 其他應收/其他應付/代收/代付款 ====

    DATE_COL = "過帳日期"
    UNCLEARED_COL = "結清文件"

    RECEIVABLE_CODES = {
        "11780100",  # 其他應收款-非聯屬公司
        "11780200",  # 其他應收款-其他
        "11880100",  # 其他應收款-聯屬公司
    }
    PAYABLE_CODES = {
        "21710100",  # 應付薪資
        "21710200",  # 應付獎金
        "21710500",  # 暫估應付薪資
        "21720100",  # 應付租金
        "21740100",  # 暫估應付費用
        "21780101",  # 應付費用-非聯屬
        "21780102",  # 應付費用-聯屬
        "21780300",  # 應付勞務
        "21900202",  # 其他應付費用-聯屬
        "22280201",  # 其他應付費用-非聯屬
    }
    COLLECTION_PAYMENT_CODES = {
        "22820100",  # 代扣稅款
        "22820200",  # 其他代收款
        "22820205",  # 其他代收款-代扣五险一金
        "12820100",  # 代付款
    }

    # Only run if columns exist
    missing_cols_90 = [c for c in [DATE_COL, UNCLEARED_COL, gl_col] if c not in df_export.columns]
    if not missing_cols_90:
        # Build or reuse df_tmp with helper columns
        if "df_tmp" not in locals():
            df_tmp = df_export.copy()
        if "_posting_date" not in df_tmp.columns:
            dt = pd.to_datetime(df_tmp[DATE_COL], errors="coerce")
            df_tmp["_posting_date"] = dt.dt.date
        if "_code" not in df_tmp.columns:
            df_tmp["_code"] = df_tmp[gl_col].apply(norm_code)

        co = cutoff_date or date.today()
        uncleared_mask = df_tmp[UNCLEARED_COL].isna() | (df_tmp[UNCLEARED_COL].astype(str).str.strip() == "")
        age90_mask = df_tmp["_posting_date"].notna() & (
            (pd.to_datetime(co) - pd.to_datetime(df_tmp["_posting_date"])).dt.days > 90
        )

        # Start appending to the same 說明 sheet: 2 blank rows after previous content
        row_ptr = ws.max_row + 2

        # Top-level title
        ws.cell(row=row_ptr, column=1, value="5. 超過90天之其他應收/其他應付/代收/代付款原因。")
        row_ptr += 2  # one blank row before first subsection

        def write_section_inline(ws_, start_row: int, section_title: str, codes: set[str]) -> int:
            """
            Writes a subsection inside 說明 at start_row.
            Returns the next row index after the subsection plus one blank line.
            """
            # Subsection title
            ws_.cell(row=start_row, column=1, value=f"— {section_title}")
            # Filter rows
            code_mask = df_tmp["_code"].isin(codes)
            sub = df_tmp[uncleared_mask & age90_mask & code_mask].copy()

            # If empty → write '無 ...'
            if sub.empty:
                ws_.cell(row=start_row + 1, column=1, value=f"無 {section_title}")
                return start_row + 2 + 1  # (title + line) then +1 blank row

            # If not empty → one blank row, then header + data
            hdr_row = start_row + 2
            for j, col_name in enumerate(selected_cols, start=1):
                ws_.cell(row=hdr_row, column=j, value=str(col_name))
            copy_header_style(src_ws, [df_export.columns.get_loc(c) + 1 for c in selected_cols], ws_, dst_row=hdr_row)

            # Data
            DATE_COLS = {"文件日期", "過帳日期"}
            header_to_idx = {h: idx + 1 for idx, h in enumerate(selected_cols)}
            r = hdr_row + 1
            for row_vals in sub[selected_cols].itertuples(index=False, name=None):
                for j, val in enumerate(row_vals, start=1):
                    header = selected_cols[j - 1]
                    v = to_date_value(val) if header in DATE_COLS else val
                    ws_.cell(row=r, column=j, value=v)
                r += 1

            # Date format
            for h in DATE_COLS:
                if h in header_to_idx:
                    cidx = header_to_idx[h]
                    for rr in range(hdr_row + 1, r):
                        ws_.cell(row=rr, column=cidx).number_format = "m/d/yyyy"


            # Highlight for all codes included in this sub-section
            for ccode, subgrp in sub.groupby("_code"):
                _highlight_code_rows(ccode, subgrp)

            # One blank row after the block
            return r + 1

        # Sections with ONE blank row between them
        row_ptr = write_section_inline(ws, row_ptr, "超過90天其他應收未沖帳明細", RECEIVABLE_CODES)
        row_ptr = write_section_inline(ws, row_ptr, "超過90天其他應付費用未沖帳明細", PAYABLE_CODES)
        row_ptr = write_section_inline(ws, row_ptr, "超過90天其他代收/代付款未沖帳明細", COLLECTION_PAYMENT_CODES)

    # ==== Step 6 (append to 說明): >30天 暫付款/暫收款 未能結清 ====
    # Reuse df_tmp, selected_cols, src_ws, etc. in current scope.

    ADVANCE_PAY_CODES = {"12810100", "12810200"}   # 暫付款
    ADVANCE_REC_CODES = {"22810100", "22810200"}   # 暫收款

    # Build 30-day mask
    age30_mask = df_tmp["_posting_date"].notna() & (
        (pd.to_datetime(co) - pd.to_datetime(df_tmp["_posting_date"])).dt.days > 30
    )

    # Start area (2 blank rows after whatever is already on 說明)
    row_ptr = ws.max_row + 2
    ws.cell(row=row_ptr, column=1, value="6. 超過30天暫付款/暫收款未能結清的合理性。")
    row_ptr += 2  # one blank row before first subsection

    def write_section_30(ws_, start_row: int, section_title: str, codes: set[str]) -> int:
        """
        Writes a subsection for 30-day 未結清 with given codes at start_row.
        Returns next row index after the block + one blank line.
        """
        ws_.cell(row=start_row, column=1, value=f"→{section_title}：")
        code_mask = df_tmp["_code"].isin(codes)
        sub = df_tmp[uncleared_mask & age30_mask & code_mask].copy()

        if sub.empty:
            ws_.cell(row=start_row + 1, column=1, value=f"無 {section_title}")
            return start_row + 2 + 1  # title + line, then 1 blank

        hdr_row = start_row + 1
        # Header
        for j, col_name in enumerate(selected_cols, start=1):
            ws_.cell(row=hdr_row, column=j, value=str(col_name))
        copy_header_style(src_ws, [df_export.columns.get_loc(c) + 1 for c in selected_cols], ws_, dst_row=hdr_row)

        # Data
        DATE_COLS = {"文件日期", "過帳日期"}
        header_to_idx = {h: idx + 1 for idx, h in enumerate(selected_cols)}
        r = hdr_row + 1
        for row_vals in sub[selected_cols].itertuples(index=False, name=None):
            for j, val in enumerate(row_vals, start=1):
                header = selected_cols[j - 1]
                v = to_date_value(val) if header in DATE_COLS else val
                ws_.cell(row=r, column=j, value=v)
            r += 1

        # Date format
        for h in DATE_COLS:
            if h in header_to_idx:
                cidx = header_to_idx[h]
                for rr in range(hdr_row + 1, r):
                    ws_.cell(row=rr, column=cidx).number_format = "m/d/yyyy"

        # Highlight for all codes in this 30-day subsection
        for ccode, subgrp in sub.groupby("_code"):
            _highlight_code_rows(ccode, subgrp)

        return r + 1  # one blank row after the block

    # Two subsections: 暫付款 / 暫收款
    row_ptr = write_section_30(ws, row_ptr, "超過30天 暫付款未沖帳明細", ADVANCE_PAY_CODES)
    row_ptr = write_section_30(ws, row_ptr, "超過30天 暫收款未沖帳明細", ADVANCE_REC_CODES)

    # === Append manual-decision questions (no answers, just text) ===
    row_ptr = ws.max_row + 2  # leave some blank space

    manual_questions = [
        "5. 關係人之應收/應付款是否逾期? 原因為何?",
        "6. 關係人交易科目的餘額是否對帳一致?",
        "7. 是否有預付/應付/應收/暫付/暫收/暫付等資產負債無檢查之情況。",
    ]

    for q in manual_questions:
        ws.cell(row=row_ptr, column=1, value=q)
        row_ptr += 2  # add a blank line after each question

    # === Format columns L and N in 說明 (if they exist) ===
    NUMBER_FMT = '#,##0;[Red](#,##0)'

    col_indexes = []
    for target_col in ["L", "N"]:
        try:
            idx = openpyxl.utils.column_index_from_string(target_col)
            if idx <= ws.max_column:  # only if column exists
                col_indexes.append(idx)
        except Exception:
            continue

    for cidx in col_indexes:
        for r in range(2, ws.max_row + 1):  # skip header row
            cell = ws.cell(row=r, column=cidx)
            # only format if numeric
            if isinstance(cell.value, (int, float)):
                cell.number_format = NUMBER_FMT

    # 7) remove original sheets in the OUTPUT (not touching your source file if you chose a new file)
    #    We will always save to output_path; if not inplace, that's a different file.
    #    Either way, we delete the listed titles if present.
    to_drop_ci = {t.lower() for t in drop_original_titles}
    for ws in list(wb.worksheets):
        if ws.title.lower() in to_drop_ci:
            del wb[ws.title]

    # 8) save
    if inplace:
        wb.save(export_path)
        saved_to = str(export_path)
    else:
        wb.save(output_path)
        saved_to = str(output_path)

    return {
        "export_sheet": sheet_used,
        "gl_col": gl_col,
        "unique_accounts": int(df_export_valid["_code"].nunique()),
        "rows_grouped": int(len(df_export_valid)),
        "saved_to": saved_to,
        "columns_B_to_X": selected_cols,
        "date_columns_formatted": sorted(date_cols_set),
        "dropped_sheets": [t for t in drop_original_titles if t.lower() in to_drop_ci],
    }


# ---------------- CLI ----------------

def main():
    p = argparse.ArgumentParser(description="Group export rows by G/L科目, create sub-sheets, format dates, and drop original sheets.")
    p.add_argument("--export", default="export-科餘-1000-asset.xlsx",
                   help="Path to the export workbook (default: ./export-科餘-1000-asset.xlsx)")
    p.add_argument("--mapping", default="會計科目對照表.xlsx",
                   help="Path to the mapping workbook (default: ./會計科目對照表.xlsx)")
    p.add_argument("--output", default=None,
                   help="Output path (ignored if --inplace). Defaults to <export>_grouped.xlsx")
    p.add_argument("--sheet", default=None,
                   help="Export sheet name to read. If omitted, uses the first sheet.")
    p.add_argument("--inplace", action="store_true",
                   help="Modify the export file in place.")
    p.add_argument("--drop-sheets", default="Sheet1,Sheet2,Sheet3",
                   help="Comma-separated sheet titles to delete from the output (case-insensitive). Default: Sheet1,Sheet2,Sheet3")
    p.add_argument("--date-cols", default="文件日期,過帳日期",
                   help="Comma-separated header names treated as dates and formatted m/d/yyyy. Default: 文件日期,過帳日期")
    
    p.add_argument("--cutoff", default=None,
               help="Cutoff date for 'older than 30 days' checks, e.g. 2025-06-30. Defaults to today.")
    
    args = p.parse_args()
    if args.cutoff:
        try:
            cutoff_date = datetime.strptime(args.cutoff, "%Y-%m-%d").date()
        except ValueError:
            print("[ERROR] --cutoff must be YYYY-MM-DD (e.g., 2025-06-30)", file=sys.stderr)
            sys.exit(2)
    else:
        cutoff_date = date.today()


    export_path = Path(args.export).expanduser().resolve()
    mapping_path = Path(args.mapping).expanduser().resolve()

    if not export_path.exists():
        print(f"[ERROR] Export file not found: {export_path}", file=sys.stderr)
        sys.exit(2)
    if not mapping_path.exists():
        print(f"[ERROR] Mapping file not found: {mapping_path}", file=sys.stderr)
        sys.exit(2)

    output_path = choose_output_path(export_path, args.inplace, args.output)
    drop_original_titles = [t.strip() for t in args.drop_sheets.split(",") if t.strip()]
    date_columns = [t.strip() for t in args.date_cols.split(",") if t.strip()]

    stats = group_export_by_account(
        export_path=export_path,
        mapping_path=mapping_path,
        output_path=output_path,
        sheet_name=args.sheet,
        inplace=args.inplace,
        cutoff_date=cutoff_date,  
        drop_original_titles=drop_original_titles,
        date_columns=date_columns,
    )

    print("[OK] Grouping complete.")
    for k, v in stats.items():
        print(f"- {k}: {v}")

if __name__ == "__main__":
    main()
