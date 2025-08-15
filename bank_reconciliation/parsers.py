import openpyxl
from pathlib import Path
import math
import pandas as pd
from utils import load_sheet
from typing import Union
from utils import is_missing_number


class BankParserBase:
    def __init__(self, path):        # path = pathlib.Path
        self.path = path

    def extract_rows(self):
        """Return list of (raw_customer_text, amount)."""
        raise NotImplementedError

# class CitiParser(BankParserBase):
#     """
#     Parses 花旗對帳單 (xls/xlsx)
#     - Prefer sheet 'Sheet2'; if missing, fall back to first sheet (index 0)
#     - Header marker: '細節描述' in column E
#     - Customer text: column E
#     - Amount: column G
#     - Start reading 2 rows below header, stop when E is blank
#     """
#     SHEET_CANDIDATES = ("Sheet2", 0)
#     CUSTOMER_COL = "E"
#     AMOUNT_COL   = "G"
#     HEADER_KEYWORD = "細節描述"

#     def _load_citi_sheet(self):
#         last_err = None
#         for candidate in self.SHEET_CANDIDATES:
#             try:
#                 return load_sheet(self.path, sheet=candidate, header=None)
#             except Exception as e:
#                 last_err = e
#         raise RuntimeError(
#             f"Could not open a valid sheet in {self.path.name} "
#             f"(tried {self.SHEET_CANDIDATES}). Last error: {last_err}"
#         )

#     def extract_rows(self):
#         sheet = self._load_citi_sheet()
#         rows = []

#         if isinstance(sheet, openpyxl.worksheet.worksheet.Worksheet):
#             # ---------- .xlsx ----------
#             ws = sheet
#             # find header rows where E == '細節描述'
#             hits = [
#                 r for r in range(1, ws.max_row + 1)
#                 if (ws[f"{self.CUSTOMER_COL}{r}"].value is not None
#                     and str(ws[f"{self.CUSTOMER_COL}{r}"].value).strip() == self.HEADER_KEYWORD)
#             ]
#             if not hits:
#                 raise RuntimeError(f"No '{self.HEADER_KEYWORD}' found in {self.path.name}")

#             hdr = hits[1] if len(hits) > 1 else hits[0]
#             r = hdr + 2

#             while r <= ws.max_row:
#                 cust = ws[f"{self.CUSTOMER_COL}{r}"].value
#                 if cust is None or not str(cust).strip():
#                     break
#                 raw_amt = ws[f"{self.AMOUNT_COL}{r}"].value
#                 amt = _to_float(raw_amt)
#                 rows.append((str(cust).strip(), amt))
#                 r += 1

#         else:
#             # ---------- .xls (DataFrame) ----------
#             df: pd.DataFrame = sheet
#             # E -> idx 4, G -> idx 6
#             header_rows = df[df[4] == self.HEADER_KEYWORD].index
#             if header_rows.empty:
#                 raise RuntimeError(f"No '{self.HEADER_KEYWORD}' found in {self.path.name}")

#             hdr = header_rows[1] if len(header_rows) > 1 else header_rows[0]
#             start = hdr + 2

#             for idx in range(start, len(df)):
#                 cust = df.at[idx, 4]
#                 if pd.isna(cust) or not str(cust).strip():
#                     break
#                 amt = _to_float(df.at[idx, 6])
#                 rows.append((str(cust).strip(), amt))

#         print(f"Loaded {len(rows)} entries from {self.path.name}")
#         return rows
class CitiParser(BankParserBase):
    """
    Parses 花旗對帳單 (xls/xlsx)
    - Prefer sheet 'Sheet2'; else first sheet
    - Start: SECOND '細節描述' row (ignore anything above)
    - Stop: FIRST '期終結餘' seen in column B
    - Keep rows where 入帳 (G) has a value; customer = E
    """
    SHEET_CANDIDATES = ("Sheet2", 0)
    CUSTOMER_COL = "E"
    AMOUNT_COL   = "G"
    HEADER_KEYWORD = "細節描述"
    STOP_TOKEN_B   = "期終結餘"

    def _load_citi_sheet(self):
        last_err = None
        for cand in self.SHEET_CANDIDATES:
            try:
                return load_sheet(self.path, sheet=cand, header=None)
            except Exception as e:
                last_err = e
        raise RuntimeError(f"Could not open a valid sheet in {self.path.name} (tried {self.SHEET_CANDIDATES}). Last error: {last_err}")

    def extract_rows(self):
        sheet = self._load_citi_sheet()
        rows = []

        if isinstance(sheet, openpyxl.worksheet.worksheet.Worksheet):
            # ---------- .xlsx ----------
            ws = sheet
            # find header rows where E == '細節描述'
            hits = [
                r for r in range(1, ws.max_row + 1)
                if (ws[f"{self.CUSTOMER_COL}{r}"].value is not None
                    and str(ws[f"{self.CUSTOMER_COL}{r}"].value).strip() == self.HEADER_KEYWORD)
            ]
            if len(hits) < 2:
                raise RuntimeError(f"Less than 2 '{self.HEADER_KEYWORD}' found in {self.path.name}")

            hdr = hits[1]          # ← SECOND occurrence
            r = hdr + 2

            while r <= ws.max_row:
                # stop at first '期終結餘' in column B
                bval = ws[f"B{r}"].value
                if bval is not None and str(bval).strip() == self.STOP_TOKEN_B:
                    break

                amt = _to_float(ws[f"{self.AMOUNT_COL}{r}"].value)
                if not amt:
                    r += 1
                    continue

                cust = ws[f"{self.CUSTOMER_COL}{r}"].value
                if cust is None or not str(cust).strip():
                    r += 1
                    continue

                rows.append((str(cust).strip(), amt))
                r += 1

        else:
            # ---------- .xls (DataFrame) ----------
            df: pd.DataFrame = sheet
            header_rows = df[df[4].astype(str).str.strip() == self.HEADER_KEYWORD].index
            if len(header_rows) < 2:
                raise RuntimeError(f"Less than 2 '{self.HEADER_KEYWORD}' found in {self.path.name}")

            hdr = int(header_rows[1])   # SECOND occurrence
            start = hdr + 2

            for idx in range(start, len(df)):
                bval = df.at[idx, 1] if 1 in df.columns else None
                if pd.notna(bval) and str(bval).strip() == self.STOP_TOKEN_B:
                    break

                amt = _to_float(df.at[idx, 6] if 6 in df.columns else None)
                if not amt:
                    continue

                cust = df.at[idx, 4] if 4 in df.columns else None
                if pd.isna(cust) or not str(cust).strip():
                    continue

                rows.append((str(cust).strip(), amt))

        print(f"Loaded {len(rows)} entries from {self.path.name}")
        return rows




class CTBCParser(BankParserBase):
    """
    Parses 1000-中信-*.xls/.xlsx
    - Header row has '備註' in column J
    - Customer name is under column J
    - Amount is in column E
    - We read the first sheet (index 0) for both xls/xlsx.
    - Stop when column J becomes blank.
    """
    SHEET = 0
    CUSTOMER_COL   = "J"         # 1-based Excel col
    AMOUNT_COL     = "E"
    HEADER_KEYWORD = "備註"

    def extract_rows(self):
        sheet = load_sheet(self.path, sheet=self.SHEET, header=None)
        rows = []

        if isinstance(sheet, openpyxl.worksheet.worksheet.Worksheet):
            # ---- .xlsx path ----
            ws = sheet

            # 1) find header row where J == '備註'
            hdr = None
            for r in range(1, ws.max_row + 1):
                val = ws[f"{self.CUSTOMER_COL}{r}"].value
                if val is not None and str(val).strip() == self.HEADER_KEYWORD:
                    hdr = r
                    break
            if hdr is None:
                raise RuntimeError(f"No '{self.HEADER_KEYWORD}' in {self.path.name}")

            # 2) read until J is blank
            r = hdr + 1
            while r <= ws.max_row:
                cust = ws[f"{self.CUSTOMER_COL}{r}"].value
                if cust is None or not str(cust).strip():
                    break

                raw_amt = ws[f"{self.AMOUNT_COL}{r}"].value
                amt = _to_float(raw_amt)

                rows.append((str(cust).strip(), amt))
                r += 1

        else:
            # ---- .xls path (pandas DataFrame) ----
            df: pd.DataFrame = sheet
            # J → idx 9, E → idx 4
            header_rows = df[df[9] == self.HEADER_KEYWORD].index
            if header_rows.empty:
                raise RuntimeError(f"No '{self.HEADER_KEYWORD}' in {self.path.name}")
            hdr = header_rows[0]

            for idx in range(hdr + 1, len(df)):
                cust = df.at[idx, 9]
                if pd.isna(cust) or not str(cust).strip():
                    break

                amt = _to_float(df.at[idx, 4])
                rows.append((str(cust).strip(), amt))

        print(f"Loaded {len(rows)} entries from {self.path.name}")
        return rows


class MegaParser(BankParserBase):
    """
    Parses 1000-兆豐-*.xls[x]
    - Header row has '存入金額' in column F
    - Customer name sits under '備註' in column H
    - Stop reading once column D contains '總計'
    """
    SHEET      = 0           # single‐sheet workbooks
    AMOUNT_COL = "F"
    CUSTOMER_COL = "H"
    STOP_COL   = "D"
    HEADER_KEYWORD = "存入金額"
    STOP_TOKEN = "總計"

    def extract_rows(self) -> list[tuple[str, float]]:
        # load_sheet will return either:
        #  - openpyxl.Worksheet for .xlsx
        #  - pandas.DataFrame   for .xls
        sheet = load_sheet(self.path, sheet=self.SHEET, header=None)

        # --- XLSX path (openpyxl) ---
        if isinstance(sheet, openpyxl.worksheet.worksheet.Worksheet):
            ws = sheet
            # 1) find header row in column F
            hdr = None
            for r in range(1, ws.max_row + 1):
                if ws[f"{self.AMOUNT_COL}{r}"].value == self.HEADER_KEYWORD:
                    hdr = r
                    break
            if hdr is None:
                raise RuntimeError(f"No '{self.HEADER_KEYWORD}' in {self.path.name}")

            # 2) walk until STOP_TOKEN in column D
            rows = []
            r = hdr + 1
            while r <= ws.max_row:
                if ws[f"{self.STOP_COL}{r}"].value == self.STOP_TOKEN:
                    break

                cust = ws[f"{self.CUSTOMER_COL}{r}"].value
                amt  = ws[f"{self.AMOUNT_COL}{r}"].value

                if cust is None or not str(cust).strip():
                    break

                rows.append((str(cust).strip(), amt))
                r += 1

            return rows

        # --- XLS path (pandas) ---
        elif isinstance(sheet, pd.DataFrame):
            df: pd.DataFrame = sheet
            # column letters to 0-based indices: D=3, F=5, H=7
            COL_D, COL_F, COL_H = 3, 5, 7

            # 1) find header row by matching HEADER_KEYWORD in column F
            mask = df[COL_F] == self.HEADER_KEYWORD
            if not mask.any():
                raise RuntimeError(f"No '{self.HEADER_KEYWORD}' in {self.path.name}")
            hdr = mask.idxmax()  # first occurrence

            # 2) iterate until STOP_TOKEN appears in column D
            rows: list[tuple[str, float]] = []
            for idx in range(hdr + 1, len(df)):
                if df.at[idx, COL_D] == self.STOP_TOKEN:
                    break

                cust = df.at[idx, COL_H]
                if pd.isna(cust) or not str(cust).strip():
                    break

                amt = df.at[idx, COL_F]
                # optionally convert comma‐thousands to float
                amt = float(amt.replace(",", "")) if isinstance(amt, str) else amt

                rows.append((cust.strip(), amt))

            return rows

        else:
            raise TypeError(f"Unexpected sheet type: {type(sheet)}")
        
def _to_float(x):
    if x is None: return None
    if isinstance(x, str):
        x = x.replace(",", "").strip()
        if x == "": return None
    try:
        return float(x)
    except Exception:
        return None


class FubonParser(BankParserBase):
    """
    Parses 1000-富邦-*.xls/.xlsx
    - Header row has '存入金額' in column F
    - Customer name sits under '附言' in column I
    - Stop reading once column A contains '小計'
    - Sheet can be named '報表' or 'Sheet1' (prefer '報表')
    """
    SHEET_CANDIDATES = ("報表", "Sheet1")   # try in order
    AMOUNT_COL     = "F"
    CUSTOMER_COL   = "I"
    HEADER_KEYWORD = "存入金額"
    STOP_TOKEN     = "小計"

    def _load_fubon_sheet(self):
        last_err = None
        for candidate in (*self.SHEET_CANDIDATES, 0):  # fallback to first sheet
            try:
                return load_sheet(self.path, sheet=candidate, header=None)
            except Exception as e:
                last_err = e
        raise RuntimeError(
            f"Could not open a valid sheet in {self.path.name} "
            f"(tried {self.SHEET_CANDIDATES} and first sheet). Last error: {last_err}"
        )

    def extract_rows(self):
        sheet = self._load_fubon_sheet()
        rows = []

        if isinstance(sheet, openpyxl.worksheet.worksheet.Worksheet):
            # ── .xlsx path
            ws = sheet
            # 1) find header row (F == '存入金額')
            hdr = None
            for r in range(1, ws.max_row + 1):
                v = ws[f"{self.AMOUNT_COL}{r}"].value
                if (isinstance(v, str) and v.strip() == self.HEADER_KEYWORD) or v == self.HEADER_KEYWORD:
                    hdr = r
                    break
            if hdr is None:
                raise RuntimeError(f"No '{self.HEADER_KEYWORD}' in {self.path.name}")

            # 2) walk down until A == '小計'
            r = hdr + 1
            while r <= ws.max_row:
                aval = ws[f"A{r}"].value
                if aval is not None and str(aval).strip() == self.STOP_TOKEN:
                    break

                # amount in F; if empty/zero → skip (do NOT read customer)
                amt = _to_float(ws[f"{self.AMOUNT_COL}{r}"].value)
                if is_missing_number(amt) or amt == 0:
                    # skip row; do NOT read/compare customer
                    r += 1
                    continue

                # customer in I; if blank → skip (but continue scanning)
                cust = ws[f"{self.CUSTOMER_COL}{r}"].value
                if cust is None or not str(cust).strip():
                    r += 1
                    continue

                rows.append((str(cust).strip(), amt))
                r += 1

        else:
            # ── .xls path via pandas DataFrame (A=0, F=5, I=8)
            df: pd.DataFrame = sheet

            # 1) header row
            header_rows = df[df[5].astype(str).str.strip() == self.HEADER_KEYWORD].index
            if header_rows.empty:
                raise RuntimeError(f"No '{self.HEADER_KEYWORD}' in {self.path.name}")
            hdr = int(header_rows[0])

            # 2) iterate until A == '小計'
            for idx in range(hdr + 1, len(df)):
                aval = df.at[idx, 0] if 0 in df.columns else None
                if pd.notna(aval) and str(aval).strip() == self.STOP_TOKEN:
                    break

                amt = _to_float(df.at[idx, 5] if 5 in df.columns else None)
                if is_missing_number(amt) or amt == 0:
                    # skip row; do NOT read/compare customer
                    continue

                cust = df.at[idx, 8] if 8 in df.columns else None
                if pd.isna(cust) or not str(cust).strip():
                    continue

                rows.append((str(cust).strip(), amt))

        print(f"Loaded {len(rows)} entries from {self.path.name}")
        return rows
    

class SinopacParser(BankParserBase):
    """
    Parses 1000-永豐-*.xls/.xlsx
    - Header row has '存入' in column F
    - Customer name sits under '備註' in column J
    - Stop when you hit a truly blank customer cell
    """
    SHEET_NAME     = "工作表1"   # or leave None to use the first sheet
    AMOUNT_COL     = "F"
    CUSTOMER_COL   = "J"
    HEADER_KEYWORD = "存入"

    def extract_rows(self):
        sheet = load_sheet(self.path,
                           sheet=self.SHEET_NAME,
                           header=None)

        rows = []
        # .xlsx path
        if isinstance(sheet, openpyxl.worksheet.worksheet.Worksheet):
            ws = sheet
            # 1) locate header row in column F
            hdr = None
            for r in range(1, ws.max_row+1):
                if ws[f"{self.AMOUNT_COL}{r}"].value == self.HEADER_KEYWORD:
                    hdr = r
                    break
            if hdr is None:
                raise RuntimeError(f"No '{self.HEADER_KEYWORD}' in {self.path.name}")

            # 2) read down until customer cell is blank
            r = hdr + 1
            while r <= ws.max_row:
                cust = ws[f"{self.CUSTOMER_COL}{r}"].value
                amt  = ws[f"{self.AMOUNT_COL}{r}"].value

                # stop on blank customer
                if cust is None or not str(cust).strip():
                    break

                cust_text = str(cust).strip()
                # normalize amt → float if it's text
                if isinstance(amt, str):
                    amt = float(amt.replace(",", ""))
                rows.append((cust_text, amt))
                r += 1

        # .xls path via pandas
        else:
            df: pd.DataFrame = sheet
            # A=0, F=5, J=9
            header_rows = df[df[5] == self.HEADER_KEYWORD].index
            if header_rows.empty:
                raise RuntimeError(f"No '{self.HEADER_KEYWORD}' in {self.path.name}")
            hdr = header_rows[0]

            for idx in range(hdr + 1, len(df)):
                cust = df.at[idx, 9]
                if pd.isna(cust) or not str(cust).strip():
                    break
                amt = df.at[idx, 5]
                if isinstance(amt, str):
                    amt = float(amt.replace(",", ""))
                rows.append((str(cust).strip(), amt))

        print(f"Loaded {len(rows)} entries from {self.path.name}")
        return rows


class ESunParser(BankParserBase):
    """
    Parses 1000-玉山-*.xls/.xlsx
    - Header row has '存' in column G
    - Deposit amount in column G
    - Customer name under '備註' in column I
    - Stop reading once column B contains '總計'
    """
    SHEET_NAME     = 0
    AMOUNT_COL     = "G"
    CUSTOMER_COL   = "I"
    HEADER_KEYWORD = "存"
    STOP_TOKEN     = "總計"

    def extract_rows(self):
        # load_sheet gives openpyxl WS or pandas DF
        sheet = load_sheet(self.path, sheet=self.SHEET_NAME, header=None)
        rows = []

        # —— .xlsx path —— 
        if isinstance(sheet, openpyxl.worksheet.worksheet.Worksheet):
            ws = sheet
            # 1) find the header row in G
            hdr = None
            for r in range(1, ws.max_row + 1):
                if ws[f"{self.AMOUNT_COL}{r}"].value == self.HEADER_KEYWORD:
                    hdr = r
                    break
            if hdr is None:
                raise RuntimeError(f"No '{self.HEADER_KEYWORD}' in {self.path.name}")

            # 2) walk down until B == '總計'
            r = hdr + 1
            while r <= ws.max_row:
                if ws[f"B{r}"].value == self.STOP_TOKEN:
                    break

                cust = ws[f"{self.CUSTOMER_COL}{r}"].value
                amt  = ws[f"{self.AMOUNT_COL}{r}"].value

                # if blank customer, stop
                if cust is None or not str(cust).strip():
                    break

                # normalize strings to floats
                if isinstance(amt, str):
                    amt = float(amt.replace(",", ""))
                rows.append((str(cust).strip(), amt))
                r += 1

        # —— .xls path —— 
        else:
            df: pd.DataFrame = sheet
            # col G -> idx 6 (0-based), col B -> idx 1, col I -> idx 8
            # 1) header row where df[6] == HEADER_KEYWORD
            hdrs = df[df[6] == self.HEADER_KEYWORD].index
            if hdrs.empty:
                raise RuntimeError(f"No '{self.HEADER_KEYWORD}' in {self.path.name}")
            hdr = hdrs[0]

            # 2) read until STOP_TOKEN in column 1
            for idx in range(hdr + 1, len(df)):
                if df.at[idx, 1] == self.STOP_TOKEN:
                    break

                cust = df.at[idx, 8]
                amt  = df.at[idx, 6]

                if pd.isna(cust) or not str(cust).strip():
                    break

                if isinstance(amt, str):
                    amt = float(amt.replace(",", ""))

                rows.append((str(cust).strip(), amt))

        print(f"Loaded {len(rows)} entries from {self.path.name}")
        return rows
