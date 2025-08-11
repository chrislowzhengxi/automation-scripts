# #!/usr/bin/env python3
# import argparse, os, re
# from pathlib import Path
# from copy import copy
# from openpyxl import load_workbook


# PERIOD_RX = re.compile(r"(20\d{2})(0[1-9]|1[0-2])")  # finds 202401..202512
# TARGET_SHEET = "2-2.銷貨倍力"  # change if your sheet name differs
# CODE_TO_CELL = {
#     "421007": "C18",  # 聯屬公司服務收入
#     "421807": "C19",  # 聯屬暫估收入
# }
# COLUMN_A = 1   # A = 1
# COLUMN_S = 19  # S = 19

# # ---------- common helpers ----------
# def find_downloads() -> Path:
#     # Windows-friendly default
#     userprofile = os.environ.get("USERPROFILE")
#     if userprofile:
#         d = Path(userprofile) / "Downloads"
#         if d.exists():
#             return d
#     # Fallback to home dir
#     return Path.home() / "Downloads"


# def pick_file_by_period(prefix: str, period: str, explicit: str | None) -> Path:
#     if explicit:
#         p = Path(explicit)
#         if not p.exists():
#             raise FileNotFoundError(f"Path not found: {p}")
#         return p

#     downloads = find_downloads()
#     candidates = sorted(Path(downloads).glob(f"*{prefix}*.xlsx"))
#     if not candidates:
#         raise FileNotFoundError(f"No {prefix} files in {downloads}")

#     # 1) exact period match
#     period_matches = [p for p in candidates if period in p.name]
#     if period_matches:
#         return max(period_matches, key=lambda p: p.stat().st_mtime)

#     # 2) else pick highest YYYYMM we can parse
#     def pnum(p: Path):
#         m = PERIOD_RX.search(p.name)
#         return int(m.group(1) + m.group(2)) if m else None

#     with_period = [(p, pnum(p)) for p in candidates]
#     with_period = [(p, n) for (p, n) in with_period if n is not None]
#     if with_period:
#         max_period = max(n for _, n in with_period)
#         newest_in_max = max([p for (p, n) in with_period if n == max_period],
#                             key=lambda p: p.stat().st_mtime)
#         return newest_in_max

#     # 3) fallback newest by mtime
#     return max(candidates, key=lambda p: p.stat().st_mtime)

# # def pick_mrs_file(period: str, explicit: str | None) -> Path:
# #     if explicit:
# #         p = Path(explicit)
# #         if not p.exists():
# #             raise FileNotFoundError(f"Explicit MRS path not found: {p}")
# #         return p

# #     downloads = find_downloads()
# #     candidates = sorted(Path(downloads).glob("*MRS0014*.xlsx"))
# #     if not candidates:
# #         raise FileNotFoundError(f"No MRS0014 files found in {downloads}")

# #     # 1) try exact period match
# #     period_matches = [p for p in candidates if period in p.name]
# #     if period_matches:
# #         return max(period_matches, key=lambda p: p.stat().st_mtime)

# #     # 2) otherwise pick the highest (YYYYMM) period we can parse
# #     def extract_period_int(p: Path) -> int | None:
# #         m = PERIOD_RX.search(p.name)
# #         if not m:
# #             return None
# #         return int(m.group(1) + m.group(2))  # e.g., 202504 -> 202504

# #     with_period = [(p, extract_period_int(p)) for p in candidates]
# #     with_period = [(p, n) for (p, n) in with_period if n is not None]
# #     if with_period:
# #         # choose the file(s) with the max period, then the newest mtime
# #         max_period = max(n for _, n in with_period)
# #         newest_in_max = max([p for (p, n) in with_period if n == max_period],
# #                             key=lambda p: p.stat().st_mtime)
# #         return newest_in_max

# #     # 3) last resort: newest by modified time
# #     return max(candidates, key=lambda p: p.stat().st_mtime)


# def copy_cell(src, dst):
#     # value
#     dst.value = src.value
#     # number format + basic style so red/parentheses/accounting look right
#     dst.number_format = src.number_format
#     dst.font = copy(src.font)
#     dst.fill = copy(src.fill)
#     dst.border = copy(src.border)
#     dst.alignment = copy(src.alignment)
#     dst.protection = copy(src.protection)


# def read_codes_from_mrs(mrs_path: Path) -> dict:
#     wb = load_workbook(mrs_path, data_only=True)
#     want = {"421007": None, "421807": None}
#     try:
#         for ws in wb.worksheets:
#             max_row = ws.max_row or 5000
#             for r in range(1, max_row + 1):
#                 a_val = ws.cell(row=r, column=COLUMN_A).value
#                 if a_val is None:
#                     continue
#                 key = str(a_val).strip()
#                 if key in want:
#                     s_val = ws.cell(row=r, column=COLUMN_S).value
#                     # if multiple matches ever appear, sum them
#                     if want[key] is None:
#                         want[key] = s_val or 0
#                     else:
#                         try:
#                             want[key] += (s_val or 0)
#                         except TypeError:
#                             # if some weird text pops up, ignore that extra
#                             pass
#             if all(v is not None for v in want.values()):
#                 break
#     finally:
#         wb.close()
#     # default missing to 0 so the sheet still fills
#     return {k: (0 if v is None else v) for k, v in want.items()}

# def write_to_template(template_path: Path, out_path: Path, values: dict):
#     wb = load_workbook(template_path)
#     try:
#         ws = wb[TARGET_SHEET] if TARGET_SHEET in wb.sheetnames else wb.active
#         # fill C18 / C19
#         for code, cell_addr in CODE_TO_CELL.items():
#             ws[cell_addr] = values.get(code, 0)
#         # C21 = SUM(C18:C19)
#         ws["C21"] = "=SUM(C18:C19)"
#         out_path.parent.mkdir(parents=True, exist_ok=True)
#         wb.save(out_path)
#     finally:
#         wb.close()

# def main():
#     ap = argparse.ArgumentParser(description="Fill 2-2.銷貨倍力 from MRS0014.")
#     ap.add_argument("--period", required=True, help="e.g. 202504")
#     ap.add_argument("--mrs", help="Path to MRS0014 file (optional). If omitted, pick newest from Downloads.")
#     ap.add_argument("--template", required=True, help="Path to the 2-2 template workbook (.xlsx).")
#     ap.add_argument("--out", help="Output path (.xlsx). If omitted, auto under data/output.")
#     args = ap.parse_args()

#     mrs_path = pick_mrs_file(args.period, args.mrs)
#     template_path = Path(args.template)
#     if not template_path.exists():
#         raise FileNotFoundError(f"Template not found: {template_path}")

#     values = read_codes_from_mrs(mrs_path)

#     out_path = Path(args.out) if args.out else Path(f"ytm_forms/data/output/2-2_{args.period}.xlsx")
#     write_to_template(template_path, out_path, values)

#     print(f"Done. Wrote C18={values['421007']} C19={values['421807']} → {out_path}")

# if __name__ == "__main__":
#     main()


#!/usr/bin/env python3
import argparse, os, re
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
from copy import copy

TARGET_SHEET = "2-2.銷貨倍力"
COLUMN_A = 1
COLUMN_S = 19
PERIOD_RX = re.compile(r"(20\d{2})(0[1-9]|1[0-2])")

# ---------- common helpers ----------
def find_downloads() -> Path:
    up = os.environ.get("USERPROFILE")
    d = Path(up) / "Downloads" if up else Path.home() / "Downloads"
    return d

def pick_file_by_period(prefix: str, period: str, explicit: str | None) -> Path:
    if explicit:
        p = Path(explicit)
        if not p.exists():
            raise FileNotFoundError(f"Path not found: {p}")
        return p

    downloads = find_downloads()
    candidates = sorted(Path(downloads).glob(f"*{prefix}*.xlsx"))
    if not candidates:
        raise FileNotFoundError(f"No {prefix} files in {downloads}")

    # 1) exact period match
    period_matches = [p for p in candidates if period in p.name]
    if period_matches:
        return max(period_matches, key=lambda p: p.stat().st_mtime)

    # 2) else pick highest YYYYMM we can parse
    def pnum(p: Path):
        m = PERIOD_RX.search(p.name)
        return int(m.group(1) + m.group(2)) if m else None

    with_period = [(p, pnum(p)) for p in candidates]
    with_period = [(p, n) for (p, n) in with_period if n is not None]
    if with_period:
        max_period = max(n for _, n in with_period)
        newest_in_max = max([p for (p, n) in with_period if n == max_period],
                            key=lambda p: p.stat().st_mtime)
        return newest_in_max

    # 3) fallback newest by mtime
    return max(candidates, key=lambda p: p.stat().st_mtime)

def copy_cell(src, dst):
    # value
    dst.value = src.value
    # number format + basic style so red/parentheses/accounting look right
    dst.number_format = src.number_format
    dst.font = copy(src.font)
    dst.fill = copy(src.fill)
    dst.border = copy(src.border)
    dst.alignment = copy(src.alignment)
    dst.protection = copy(src.protection)

def dest_anchor(ws, r, c):
    for rng in ws.merged_cells.ranges:
        if rng.min_row <= r <= rng.max_row and rng.min_col <= c <= rng.max_col:
            return rng.min_row, rng.min_col
    return r, c

# ---------- task: MRS0014 → 2-2 ----------
def run_mrs0014(template_path: Path, period: str, mrs_path: str | None, out_path: Path):
    mrs_file = pick_file_by_period("MRS0014", period, mrs_path)

    # read values for 421007 / 421807 from column S
    wb_src = load_workbook(mrs_file, data_only=True)
    vals = {"421007": 0, "421807": 0}
    try:
        for ws in wb_src.worksheets:
            for r in range(1, (ws.max_row or 5000) + 1):
                a_val = ws.cell(row=r, column=COLUMN_A).value
                if a_val is None:
                    continue
                key = str(a_val).strip()
                if key in vals:
                    s_val = ws.cell(row=r, column=COLUMN_S).value or 0
                    try:
                        vals[key] += s_val
                    except TypeError:
                        pass
    finally:
        wb_src.close()

    # write to template (C18, C19, sum to C21)
    wb = load_workbook(template_path)
    try:
        ws = wb[TARGET_SHEET] if TARGET_SHEET in wb.sheetnames else wb.active
        ws["C18"] = vals["421007"]
        ws["C19"] = vals["421807"]
        ws["C21"] = "=SUM(C18:C19)"
        out_path.parent.mkdir(parents=True, exist_ok=True)
        wb.save(out_path)
    finally:
        wb.close()

# ---------- task: RPTIS10 → 2-2 ----------
def run_rptis10(template_path: Path, period: str, rpt_path: str | None, out_path: Path,
                src_sheet: str | None = None, src_rows=(6,11), src_cols=(1,3),
                dst_start_row=7, dst_start_col=1):
    rpt_file = pick_file_by_period("RPTIS10", period, rpt_path)

    wb_src = load_workbook(rpt_file, data_only=False)  # keep styles
    try:
        ws_src = wb_src[src_sheet] if (src_sheet and src_sheet in wb_src.sheetnames) else wb_src.active
        r1, r2 = src_rows
        c1, c2 = src_cols
        block = []
        for r in range(r1, r2+1):
            row_cells = []
            for c in range(c1, c2+1):
                row_cells.append(ws_src.cell(row=r, column=c))
            block.append(row_cells)
    finally:
        wb_src.close()

    wb = load_workbook(template_path)
    try:
        ws_dst = wb[TARGET_SHEET] if TARGET_SHEET in wb.sheetnames else wb.active
        # for i, row_cells in enumerate(block):
        #     dst_row = dst_start_row + i
        #     for j, src_cell in enumerate(row_cells):
        #         dst_col = dst_start_col + j
        #         copy_cell(src_cell, ws_dst.cell(row=dst_row, column=dst_col))
        anchors_written = set()
        for i, row_cells in enumerate(block):
            dst_row = dst_start_row + i
            for j, src_cell in enumerate(row_cells):
                dst_col = dst_start_col + j
                ar, ac = dest_anchor(ws_dst, dst_row, dst_col)
                key = (ar, ac)
                if key in anchors_written:
                    continue
                anchors_written.add(key)
                dst_cell = ws_dst.cell(row=ar, column=ac)
                copy_cell(src_cell, dst_cell)
        out_path.parent.mkdir(parents=True, exist_ok=True)
        wb.save(out_path)
    finally:
        wb.close()

# ---------- CLI ----------
def main():
    ap = argparse.ArgumentParser(description="Fill YTM forms")
    ap.add_argument("--task", required=True, choices=["mrs0014", "rptis10"])
    ap.add_argument("--period", required=True, help="e.g., 202504")
    ap.add_argument("--template", required=True, help="Path to the template workbook to fill")
    ap.add_argument("--out", help="Output path (.xlsx). Omit and use --inplace to overwrite template")
    ap.add_argument("--inplace", action="store_true", help="Overwrite the template file")
    ap.add_argument("--mrs", help="Explicit path to MRS0014 (optional)")
    ap.add_argument("--rptis", help="Explicit path to RPTIS10 (optional)")
    ap.add_argument("--rpt-source-sheet", help="RPTIS10 source sheet name (optional)")
    args = ap.parse_args()

    template_path = Path(args.template)
    if not template_path.exists():
        raise FileNotFoundError(f"Template not found: {template_path}")

    out_path = Path(args.out) if args.out else (template_path if args.inplace else Path(f"ytm_forms/data/output/{args.task}_{args.period}.xlsx"))

    if args.task == "mrs0014":
        run_mrs0014(template_path, args.period, args.mrs, out_path)
    else:
        run_rptis10(template_path, args.period, args.rptis, out_path, src_sheet=args.rpt_source_sheet)

    print(f"Done → {out_path}")

if __name__ == "__main__":
    main()

