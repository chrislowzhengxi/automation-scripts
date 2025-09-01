#!/usr/bin/env python3
import argparse, os, shutil
from datetime import datetime
from pathlib import Path
from copy import copy
from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string, get_column_letter
import re
from datetime import datetime
from copy import copy as copy_style
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.cell_range import CellRange

# Base = ytm_forms/
BASE_DIR = Path(__file__).resolve().parents[1]
DEFAULT_OUT_DIR = BASE_DIR / "data" / "output"

PROJECT_ROOT = Path(__file__).resolve().parents[2]
OUTPUT_DIR = PROJECT_ROOT / "ytm_forms" / "data" / "output"

# ---------- helpers ----------
def ensure_parent(p: Path):
    p.parent.mkdir(parents=True, exist_ok=True)

def copy_cell_full(src_cell, dst_cell):
    dst_cell.value = src_cell.value
    dst_cell.number_format = src_cell.number_format
    dst_cell.font = copy(src_cell.font)
    dst_cell.fill = copy(src_cell.fill)
    dst_cell.border = copy(src_cell.border)
    dst_cell.alignment = copy(src_cell.alignment)
    dst_cell.protection = copy(src_cell.protection)

def copy_col_widths(ws_src, ws_dst, c1, c2, dst_start_col):
    for j in range(c2 - c1 + 1):
        src_col_idx = c1 + j
        dst_col_idx = dst_start_col + j
        src_letter = get_column_letter(src_col_idx)
        dst_letter = get_column_letter(dst_col_idx)
        try:
            ws_dst.column_dimensions[dst_letter].width = ws_src.column_dimensions[src_letter].width
        except Exception:
            pass

def copy_row_heights(ws_src, ws_dst, r1, r2, dst_start_row):
    for i in range(r2 - r1 + 1):
        src_row = r1 + i
        dst_row = dst_start_row + i
        try:
            ws_dst.row_dimensions[dst_row].height = ws_src.row_dimensions[src_row].height
        except Exception:
            pass

def clear_sheet(ws):
    ws.delete_rows(1, ws.max_row if ws.max_row else 1)

def copy_block(ws_src, ws_dst, r1, r2, c1, c2, dst_row, dst_col):
    # copy values/styles
    for i in range(r2 - r1 + 1):
        for j in range(c2 - c1 + 1):
            s = ws_src.cell(row=r1 + i, column=c1 + j)
            d = ws_dst.cell(row=dst_row + i, column=dst_col + j)
            copy_cell_full(s, d)
    # aesthetics
    copy_row_heights(ws_src, ws_dst, r1, r2, dst_row)
    copy_col_widths(ws_src, ws_dst, c1, c2, dst_col)

def load_output_from_template(template_path: Path, out_path: Path, inplace: bool) -> Path:
    if inplace:
        return template_path
    ensure_parent(out_path)
    shutil.copy(template_path, out_path)
    return out_path


def quote_sheet(name: str) -> str:
    # always quote sheet names for safety
    return f"'{name}'" if not (name.startswith("'") and name.endswith("'")) else name

def build_ext_vlookup(path: Path, sheet: str, table_range: str, key_ref: str, col_index: int) -> str:
    # 'C:\...\[file.xls]Sheet'!$B:$C  → VLOOKUP(key, that, col_index, FALSE)
    book = f"[{path.name}]"
    sheet_quoted = quote_sheet(sheet)
    xref = f"'{str(path.parent)}\\{book}{sheet_quoted[1:-1]}'!{table_range}"
    return f"VLOOKUP({key_ref},{xref},{col_index},FALSE)"

def copy_header_style(ws, src_col_idx: int, dst_col_idx: int):
    s = ws.cell(row=1, column=src_col_idx)
    d = ws.cell(row=1, column=dst_col_idx)
    d.value = s.value  # temp; caller will overwrite with new header text
    d.number_format = s.number_format
    from copy import copy as _cpy
    d.font = _cpy(s.font); d.fill = _cpy(s.fill); d.border = _cpy(s.border)
    d.alignment = _cpy(s.alignment); d.protection = _cpy(s.protection)
    # copy column width
    try:
        from openpyxl.utils import get_column_letter
        dst_letter = get_column_letter(dst_col_idx)
        src_letter = get_column_letter(src_col_idx)
        ws.column_dimensions[dst_letter].width = ws.column_dimensions[src_letter].width
    except Exception:
        pass

def copy_body_style_from_left(ws, row: int, col_idx: int):
    # style like the immediate left neighbor (common pattern in your sheets)
    if col_idx <= 1: 
        return
    s = ws.cell(row=row, column=col_idx - 1)
    d = ws.cell(row=row, column=col_idx)
    from copy import copy as _cpy
    d.number_format = s.number_format
    d.font = _cpy(s.font); d.fill = _cpy(s.fill); d.border = _cpy(s.border)
    d.alignment = _cpy(s.alignment); d.protection = _cpy(s.protection)


def _find_row_ws(ws, col_idx: int, text: str) -> int | None:
    """Find first row where ws[row, col_idx] == text (trimmed)."""
    for r in range(1, (ws.max_row or 1) + 1):
        v = ws.cell(r, col_idx).value
        if isinstance(v, str) and v.strip() == text:
            return r
    return None

def _first_company_row(ws, start_r: int) -> int | None:
    """From start_r downward, find first row that looks like a company row (A has an ID; B not '合計')."""
    r = start_r
    while r <= (ws.max_row or 1):
        a = ws.cell(r, 1).value
        b = str(ws.cell(r, 2).value or "").strip()
        if b == "合計":
            return None
        # treat any non-empty A as a data row
        if a not in (None, "", " "):
            return r
        r += 1
    return None

def _first_total_row(ws, start_r: int) -> int | None:
    """Find first row at/after start_r where column B == '合計'."""
    r = start_r
    while r <= (ws.max_row or 1):
        if str(ws.cell(r, 2).value or "").strip() == "合計":
            return r
        r += 1
    return None

def _first_sheet_name(xlsx_path: Path, fallback: str = "Sheet0") -> str:
    try:
        _wb = load_workbook(xlsx_path, read_only=True, data_only=True)
        name = _wb.sheetnames[0] if _wb.sheetnames else fallback
        _wb.close()
        return name
    except Exception:
        return fallback

def _col_ref_local(sheet_name: str, col_letter: str) -> str:
    col = col_letter.upper()
    return f"'{sheet_name}'!${col}:${col}"

def _col_ref_external(p: Path, sheet: str, col_letter: str) -> str:
    # absolute, full column
    win_dir = str(p.parent).replace("/", "\\")
    col = col_letter.upper()
    return f"'{win_dir}\\[{p.name}]{sheet}'!${col}:${col}"


def fill_yingshoukuan_block(ws, mrs0014_path: Path, mrs0014_sheet: str):
    """
    應收款 block:
      - Locate header row (col B == '應收款')
      - Company rows: first data row after the header band to the FIRST '合計' - 1
      - C = D - M
      - D = SUMIF('4-3.應收關係人科餘'!J:J, A_r, '4-3.應收關係人科餘'!Y:Y)
      - E = D / ( SUMIF(MRS0014!A:A,"112000",MRS0014!S:S)
                + SUMIF(MRS0014!A:A,"112300",MRS0014!S:S)
                + SUMIF(MRS0014!A:A,"112337",MRS0014!S:S) )
      - Totals only on the FIRST '合計' row.
    """
    hdr = _find_row_ws(ws, 2, "應收款")
    if not hdr:
        return

    # Skip the blue section title row and the field header row → data starts at hdr+2
    first_candidate = hdr + 2
    start = _first_company_row(ws, first_candidate)
    if not start:
        return

    total_row = _first_total_row(ws, start)
    if not total_row or total_row <= start:
        return

    # Prefer local sheet 'MRS0014' if present; else external workbook
    if "MRS0014" in ws.parent.sheetnames:
        rngA = _col_ref_local("MRS0014", "A")  # 'MRS0014'!$A:$A
        rngS = _col_ref_local("MRS0014", "S")  # 'MRS0014'!$S:$S
    else:
        rngA = _col_ref_external(mrs0014_path, mrs0014_sheet, "A")
        rngS = _col_ref_external(mrs0014_path, mrs0014_sheet, "S")

    # Fill detail rows
    for r in range(start, total_row):
        ws.cell(r, 3).value = f"=D{r}-M{r}"
        ws.cell(r, 4).value = f"=SUMIF('4-3.應收關係人科餘'!J:J,A{r},'4-3.應收關係人科餘'!Y:Y)"
        ws.cell(r, 5).value = (
            f"=D{r}/("
            f"SUMIF({rngA},\"112000\",{rngS})+"
            f"SUMIF({rngA},\"112300\",{rngS})+"
            f"SUMIF({rngA},\"112337\",{rngS})"
            f")"
        )

    # Totals (FIRST 合計 only)
    ws.cell(total_row, 3).value = f"=SUM(C{start}:C{total_row-1})"
    ws.cell(total_row, 4).value = f"=SUM(D{start}:D{total_row-1})"
    ws.cell(total_row, 5).value = f"=SUM(E{start}:E{total_row-1})"

    # Percent format for E
    PCT_FMT = '0.00%;-0.00%;"-"'

    for r in range(start, total_row + 1):
        ws.cell(r, 5).number_format = PCT_FMT

BLOCK_W = 9            # A–I
SPACER_COL = "J"       # width = 1
SHEET_NAME = "1-1.公告(元)"

def _find_month_header_col(ws, header_row=1):
    """Find the leftmost cell in header_row that looks like YYYY/MM."""
    for c in range(1, ws.max_column + 1):
        v = ws.cell(header_row, c).value
        if isinstance(v, str) and re.fullmatch(r"\d{4}/\d{2}", v):
            return c
    raise RuntimeError("Couldn't locate a YYYY/MM header in row 1.")

def _ym_from_period(yyyymm: str) -> str:
    return f"{yyyymm[:4]}/{yyyymm[4:]}"

def _prev_ym(ym: str) -> str:
    y, m = map(int, ym.split("/"))
    y2, m2 = (y - 1, 12) if m == 1 else (y, m - 1)
    return f"{y2}/{m2:02d}"

def prepare_month_structure(wb_or_path, sheet_name=SHEET_NAME, period_yyyymm: str = None, rptis10_path: Path = None, mrs0034_path: Path = None, mrs0014_path: Path = None):
    """
    Structure-only (safe & period-driven):
      - new_month = --period (YYYY/MM)
      - prev_month = month before --period
      - If K1 has a YYYY/MM, it must equal prev_month (subsequent runs) → snapshot K–S
      - Else (first run), B1 must equal prev_month → snapshot A–I and ALSO write that snapshot to K–S
      - Write snapshot to A–I (new month block)
      - Set J as thin spacer and set B1 = new_month
    """
    if not period_yyyymm:
        raise ValueError("prepare_month_structure requires period_yyyymm (e.g., '202504').")

    new_month = _ym_from_period(period_yyyymm)  # "2025/04"
    prev_month = _prev_ym(new_month)            # "2025/03"

    wb = load_workbook(wb_or_path) if isinstance(wb_or_path, (str, Path)) else wb_or_path
    ws = wb[sheet_name]

    mrs0034_sheet = _first_sheet_name(mrs0034_path) if mrs0034_path else "Sheet0"
    mrs0014_sheet = _first_sheet_name(mrs0014_path) if mrs0014_path else "Sheet0"

    max_row = ws.max_row
    width   = BLOCK_W  # 9

    B1 = ws.cell(1, 2).value    # month header (expected prev_month or new_month)
    K1 = ws.cell(1, 11).value   # col K = 11 (if already shifted before)

    # Decide where the "previous month" currently lives and sanity-check it.
    if isinstance(K1, str) and re.fullmatch(r"\d{4}/\d{2}", K1):
        if K1 != prev_month:
            raise ValueError(f"[安全保護] 期望 K1='{prev_month}', 但目前是 '{K1}'.")
        prev_col = 11  # K
        need_write_prev_to_K = False
    else:
        if B1 != prev_month:
            raise ValueError(f"[安全保護] 期望 B1='{prev_month}', 但目前是 '{B1}'.")
        prev_col = 1   # A
        need_write_prev_to_K = True

    # Snapshot function
    def snapshot(col0):
        data, merges, col_widths, row_heights = [], [], [], {}
        for r in range(1, max_row + 1):
            row = []
            for i in range(width):
                c = ws.cell(r, col0 + i)
                row.append((c.value, c.font, c.fill, c.border, c.alignment,
                            c.number_format, c.protection))
            data.append(row)
        for mr in list(ws.merged_cells.ranges):
            c1, r1, c2, r2 = mr.bounds
            if col0 <= c1 and c2 <= col0 + width - 1:
                merges.append((r1, r2, c1 - col0, c2 - col0))
        for i in range(width):
            letter = get_column_letter(col0 + i)
            dim = ws.column_dimensions.get(letter)
            col_widths.append(dim.width if dim and dim.width is not None else None)
        for r, dim in ws.row_dimensions.items():
            if 1 <= r <= max_row and dim.height is not None:
                row_heights[r] = dim.height
        return (data, merges, col_widths, row_heights)

    # Restore function
    def restore(col0, snap):
        data, merges, col_widths, row_heights = snap
        for r in range(1, max_row + 1):
            for i in range(width):
                v, fnt, fill, brd, alg, numfmt, prot = data[r - 1][i]
                dc = ws.cell(r, col0 + i)
                dc.value = v
                dc.font = copy_style(fnt); dc.fill = copy_style(fill)
                dc.border = copy_style(brd); dc.alignment = copy_style(alg)
                dc.number_format = numfmt; dc.protection = copy_style(prot)
        for i, w in enumerate(col_widths):
            if w is not None:
                letter = get_column_letter(col0 + i)
                ws.column_dimensions[letter].width = w
        for r, h in row_heights.items():
            ws.row_dimensions[r].height = h
        for r1, r2, off1, off2 in merges:
            ws.merge_cells(start_row=r1, end_row=r2,
                           start_column=col0 + off1, end_column=col0 + off2)

    snap = snapshot(prev_col)

    # Write new month block A–I
    restore(1, snap)

    # Clear spacer J
    for rng in list(ws.merged_cells.ranges):
        c1, r1, c2, r2 = rng.bounds
        if c1 <= 10 <= c2:  # col J
            ws.unmerge_cells(start_row=r1, end_row=r2,
                             start_column=c1, end_column=c2)
    ws.column_dimensions[SPACER_COL].width = 1.0
    for r in range(1, max_row + 1):
        ws.cell(r, 10).value = None  # clear J

    # On first run, also materialize K–S
    if need_write_prev_to_K:
        restore(11, snap)

    # Finally set B1 to the new month
    ws.cell(1, 2).value = new_month

    # --- Delete column I (was redundant before we aligned block width to 9) ---
    ws.delete_cols(9)
    

    for col in range(1, ws.max_column + 1):
        max_length = 0
        col_letter = get_column_letter(col)
        for row in range(1, max_row + 1):
            val = ws.cell(row, col).value
            if val is not None:
                val = str(val)
                if len(val) > max_length:
                    max_length = len(val)
        # add a little padding
        ws.column_dimensions[col_letter].width = max_length + 2

    # --- Column C formulas (差異值) for new month block (A–H) ---
    row = 4
    start_row = None
    while True:
        a_val = ws.cell(row, 1).value   # company ID
        b_val = ws.cell(row, 2).value   # company name or 合計

        if b_val is None:
            break  # stop if blank row (safety)

        if str(b_val).strip() == "合計":
            # write SUM formula for the total row
            if start_row is not None and row > start_row:
                ws.cell(row, 3).value = f"=SUM(C{start_row}:C{row-1})"
            break

        # normal company row → E{r} - N{r}
        ws.cell(row, 3).value = f"=E{row}-N{row}"
        if start_row is None:
            start_row = row
        row += 1


    def _ext_ref_cell(p: Path, sheet: str, cell: str) -> str:
        r"""
        Build Excel external reference like:
        ='C:\dir\[Workbook.xlsx]SheetName'!$B$9
        Uses absolute path so links resolve regardless of output location.
        """
        import re
        m = re.match(r"([A-Za-z]+)(\d+)", cell)
        if not m:
            raise ValueError(f"Invalid cell ref: {cell}")
        col, row = m.groups()
        abs_cell = f"${col.upper()}${row}"
        win_dir = str(p.parent).replace("/", "\\")
        return f"'{win_dir}\\[{p.name}]{sheet}'!{abs_cell}"
    
    def _ext_ref_col(p: Path, sheet: str, col_letter: str) -> str:
        r"""
        Build Excel external reference for a full column:
        ='C:\dir\[Workbook.xlsx]SheetName'!$F:$F
        """
        win_dir = str(p.parent).replace("/", "\\")
        col = col_letter.upper()
        return f"'{win_dir}\\[{p.name}]{sheet}'!${col}:${col}"



    def _set_formula(ws, r: int, c: int, formula: str, debug_once: dict):
        """
        Assign formula to ws[r,c]. Strips any accidental trailing letters (e.g. '$B$9D').
        Prints the first formula set for quick inspection.
        """
        # guard: if we ever see '$B$9' or '$E$9' followed by stray letters, trim them
        for anchor in ("$B$9", "$E$9"):
            i = formula.find(anchor)
            if i != -1:
                j = i + len(anchor)
                # if anything alphabetic immediately follows, trim it
                while j < len(formula) and formula[j].isalpha():
                    j += 1
                formula = formula[:j] + formula[j:].lstrip()  # just in case of spaces

        cell = ws.cell(row=r, column=c)
        cell.value = formula

        # one-time debug print
        key = f"{get_column_letter(c)}{r}"
        if key not in debug_once:
            print(f"[formula] {key} = {formula}")
            debug_once[key] = True


    # =========================
    # Column D formulas: D = C / (external RPTIS10!$B$9)
    # =========================
    if rptis10_path is None:
        raise ValueError("rptis10_path not provided")

    # find the sheet name in the external workbook (use the first tab)
    try:
        _ext_wb = load_workbook(rptis10_path, read_only=True, data_only=True)
        ext_sheet = _ext_wb.sheetnames[0]  # e.g., 'Sheet0'
        _ext_wb.close()
    except Exception:
        # If we can't open it (permissions/locked/etc.), fall back to a sane default
        ext_sheet = "Sheet0"

    debug_once = {}

    # ----- Column D -----
    ext_b9 = _ext_ref_cell(rptis10_path, ext_sheet, "B9")
    row = 4
    start_row_d = None
    while True:
        b_val = ws.cell(row, 2).value
        if b_val is None:
            break
        if str(b_val).strip() == "合計":
            if start_row_d is not None and row > start_row_d:
                _set_formula(ws, row, 4, f"=SUM(D{start_row_d}:D{row-1})", debug_once)
            break

        _set_formula(ws, row, 4, f"=C{row}/{ext_b9}", debug_once)

        if start_row_d is None:
            start_row_d = row
        row += 1

    # =========================
    # Column E formulas:
    #  E = -SUMIF('2-3.銷貨明細'!AJ:AJ, '<this sheet>'!A{r}, '2-3.銷貨明細'!AL:AL)
    #  Stop at first 「合計」; on that row set SUM of E above.
    # =========================
    row = 4
    start_row_e = None
    total_row_e = None  # NEW: remember where 合計 is
    while True:
        b_val = ws.cell(row, 2).value  # col B
        if b_val is None:
            break
        if str(b_val).strip() == "合計":
            total_row_e = row
            if start_row_e is not None and row > start_row_e:
                ws.cell(row, 5).value = f"=SUM(E{start_row_e}:E{row-1})"
            break

        # E_r = -SUMIF('2-3.銷貨明細'!AJ:AJ, '<this sheet>'!A{r}, '2-3.銷貨明細'!AL:AL)
        ws.cell(row, 5).value = (
            f"=-SUMIF('2-3.銷貨明細'!AJ:AJ,'{sheet_name}'!A{row},'2-3.銷貨明細'!AL:AL)"
        )

        if start_row_e is None:
            start_row_e = row
        row += 1


    # =========================
    # Post-total adjustment in Column E (row just below 合計)
    # E{total+1} = E{total} + SUMIF(MRS0014!A:A, "421007", MRS0014!S:S) + SUMIF(... "421807", ...)
    # =========================
    # if total_row_e and mrs0014_path:
    #     ext_A14 = _ext_ref_col(mrs0014_path, mrs0014_sheet, "A")
    #     ext_S14 = _ext_ref_col(mrs0014_path, mrs0014_sheet, "S")
    #     target_row = total_row_e + 1
    #     copy_body_style_from_left(ws, target_row, 5)
    #     ws.cell(row=target_row, column=5).number_format = ws.cell(row=total_row_e, column=5).number_format
    #     ws.cell(row=target_row, column=5).value = (
    #         f"=E{total_row_e}"
    #         f"+SUMIF({ext_A14},\"421007\",{ext_S14})"
    #         f"+SUMIF({ext_A14},\"421807\",{ext_S14})"
    #     )
        # ----- Post-total adjustment in Column E (row just below 合計) -----
    if total_row_e and mrs0014_path:
        target_row = total_row_e + 1

        # If the next row is another section header (e.g., 銷貨/進貨/應收款/應付款), insert a spare row
        next_label = str(ws.cell(target_row, 2).value or "").strip()
        if next_label in {"銷貨", "進貨", "應收款", "應付款"}:
            ws.insert_rows(target_row)

        # Prefer local 'MRS0014' sheet; fallback to external
        if "MRS0014" in ws.parent.sheetnames:
            A14 = f"'MRS0014'!$A:$A"
            S14 = f"'MRS0014'!$S:$S"
        else:
            A14 = _ext_ref_col(mrs0014_path, mrs0014_sheet, "A")
            S14 = _ext_ref_col(mrs0014_path, mrs0014_sheet, "S")

        # Style & number format like the total row above
        copy_body_style_from_left(ws, target_row, 5)
        ws.cell(target_row, 5).number_format = ws.cell(total_row_e, 5).number_format

        # E{target} = E{total} + SUMIF(...421007...) + SUMIF(...421807...)
        ws.cell(target_row, 5).value = (
            f"=E{total_row_e}"
            f"+SUMIF({A14},\"421007\",{S14})"
            f"+SUMIF({A14},\"421807\",{S14})"
        )




    # =========================
    # ----- Column F -----
    ext_e9 = _ext_ref_cell(rptis10_path, ext_sheet, "E9")
    row = 4
    start_row_f = None
    while True:
        b_val = ws.cell(row, 2).value
        if b_val is None:
            break
        if str(b_val).strip() == "合計":
            if start_row_f is not None and row > start_row_f:
                _set_formula(ws, row, 6, f"=SUM(F{start_row_f}:F{row-1})", debug_once)
            break

        _set_formula(ws, row, 6, f"=E{row}/{ext_e9}", debug_once)

        if start_row_f is None:
            start_row_f = row
        row += 1

    # =========================
    # ----- Column G (new) -----
    # G_r = E{r} - ( SUMIFS(MRS[F:F], MRS[A:A],"421007", MRS[D:D], A{r})
    #              + SUMIFS(MRS[F:F], MRS[A:A],"421807", MRS[D:D], A{r}) )
    # Skip "合計" row (leave blank).
    # =========================
    if mrs0034_path is None:
        raise ValueError("mrs0034_path not provided")
    
    # try:
    #     _mrs_wb = load_workbook(mrs0034_path, read_only=True, data_only=True)
    #     mrs_sheet = _mrs_wb.sheetnames[0]  # e.g., 'Sheet0'
    #     _mrs_wb.close()
    # except Exception:
    #     mrs_sheet = "Sheet0"

    ext_F = _ext_ref_col(mrs0034_path, mrs0034_sheet, "F")
    ext_A = _ext_ref_col(mrs0034_path, mrs0034_sheet, "A")
    ext_D = _ext_ref_col(mrs0034_path, mrs0034_sheet, "D")

    row = 4
    while True:
        b_val = ws.cell(row, 2).value  # company name / 合計
        if b_val is None:
            break
        if str(b_val).strip() == "合計":
            # per requirement: leave G total row blank (no formula)
            break

        # Style like F’s left neighbor
        copy_body_style_from_left(ws, row, 7)  # col 7 = G

        # Build the formula
        # =E{r}-(SUMIFS(MRS[F:F],MRS[A:A],"421007",MRS[D:D],A{r})
        #       +SUMIFS(MRS[F:F],MRS[A:A],"421807",MRS[D:D],A{r}))
        # formula_g = (
        #     f"=E{row}-("
        #     f"SUMIFS({ext_F},{ext_A},\"421007\",{ext_D},A{row})+"
        #     f"SUMIFS({ext_F},{ext_A},\"421807\",{ext_D},A{row})"
        #     f")"
        # )
        formula_g = (
            f"=E{row}-("
            f"SUMPRODUCT(({ext_F})*({ext_A}=\"421007\")*({ext_D}=A{row})) + "
            f"SUMPRODUCT(({ext_F})*({ext_A}=\"421807\")*({ext_D}=A{row}))"
            f")"
        )
        _set_formula(ws, row, 7, formula_g, debug_once)  # G column
        row += 1

    CHECK_FMT = '#,##0;-#,##0;"-"'

    last_row_g = row - 1  # row has already advanced past 合計
    for r in range(4, last_row_g + 1):
        ws.cell(row=r, column=7).number_format = CHECK_FMT


    if mrs0014_path:
        fill_yingshoukuan_block(ws, mrs0014_path, mrs0014_sheet)

    return wb


# ---------- tasks ----------
def copy_43(template_path: Path, src_43: Path, out_path: Path, sheet_name="4-3.應收關係人科餘"):
    # Source: columns B..X → Destination: paste at B1 (keep alignment with template)
    c1, c2 = column_index_from_string("B"), column_index_from_string("X")
    dst_col = column_index_from_string("A")
    dst_row = 1

    wb_src = load_workbook(src_43, data_only=False)
    ws_src = wb_src.active
    r1, r2 = 1, ws_src.max_row or 1

    wb_dst = load_workbook(out_path)
    ws_dst = wb_dst[sheet_name] if sheet_name in wb_dst.sheetnames else wb_dst.active

    clear_sheet(ws_dst)  # as you said, we start with empty 4-3 each run
    copy_block(ws_src, ws_dst, r1, r2, c1, c2, dst_row, dst_col)

    wb_dst.save(out_path)
    wb_dst.close()
    wb_src.close()

def copy_23(template_path: Path, src_23: Path, out_path: Path, sheet_name="2-3.銷貨明細"):
    # Source: columns A..AJ → Destination: paste at A1
    c1, c2 = column_index_from_string("A"), column_index_from_string("AJ")
    dst_col = column_index_from_string("A")
    dst_row = 1

    wb_src = load_workbook(src_23, data_only=False)
    ws_src = wb_src.active
    r1, r2 = 1, ws_src.max_row or 1

    wb_dst = load_workbook(out_path)
    ws_dst = wb_dst[sheet_name] if sheet_name in wb_dst.sheetnames else wb_dst.active

    clear_sheet(ws_dst)  # start clean for 2-3 as well
    copy_block(ws_src, ws_dst, r1, r2, c1, c2, dst_row, dst_col)

    wb_dst.save(out_path)
    wb_dst.close()
    wb_src.close()

# --- related-party mapping (read .xls) ---
def normalize_id(x) -> str:
    if x is None:
        return ""
    s = str(x).strip()
    # keep leading zeros; if it looks like a float-ish "960286.0", strip trailing .0
    if s.endswith(".0") and s.replace(".", "", 1).isdigit():
        s = s[:-2]
    return s

def load_relparty_map(xls_path: Path) -> dict[str, str]:
    """
    Read first sheet of 關係企業(人).xls
    Col A = ID, Col C = Name. Returns {ID -> Name}.
    Requires pandas + xlrd (pip install pandas xlrd)
    """
    import pandas as pd
    df = pd.read_excel(xls_path, sheet_name=0, header=None, engine="xlrd")  # .xls
    # guard for short sheets
    if df.shape[1] < 3:
        return {}
    ids = df.iloc[:, 0].map(normalize_id)
    names = df.iloc[:, 2].fillna("").astype(str).str.strip()
    return {i: n for i, n in zip(ids, names) if i}

# Accounting format: 2-3 and 4-3
ACCOUNTING_FMT = "#,##0;[Red](#,##0);0;@"

def apply_accounting_format(ws, col_idx: int, start_row: int, end_row: int):
    for r in range(start_row, end_row + 1):
        ws.cell(row=r, column=col_idx).number_format = ACCOUNTING_FMT

# Appenders: 
def last_data_row(ws, key_col=1, max_gap=100):
    # scan down key_col (default A) until long empty gap
    rmax = ws.max_row or 1
    seen = 0
    for r in range(1, rmax + 1):
        v = ws.cell(row=r, column=key_col).value
        if v is not None and str(v).strip() != "":
            seen = r
    # if nothing found, still return 1 so headers can exist
    return max(seen, 1)


def append_calc_columns_23(ws, period: str, rates_path: Path, relparty_map: dict):
    from openpyxl.utils import column_index_from_string as colidx
    AK, AL, AM = colidx("AK"), colidx("AL"), colidx("AM")
    N, M, AJ = colidx("N"), colidx("M"), colidx("AJ")

    # headers
    copy_header_style(ws, AJ, AK); ws.cell(row=1, column=AK).value = "匯率"
    copy_header_style(ws, AJ, AL); ws.cell(row=1, column=AL).value = "換算台幣"
    copy_header_style(ws, AJ, AM); ws.cell(row=1, column=AM).value = "關係企業名稱"

    lr = last_data_row(ws, key_col=1)

    # formulas for rate / amount
    rates_vlk = lambda r: f"IF({ws.cell(row=r, column=N).coordinate}=\"NTD\",1," + \
        build_ext_vlookup(rates_path, "Summary", "$B:$C", ws.cell(row=r, column=N).coordinate, 2) + ")"

    for r in range(2, lr + 1):
        # 匯率 (AK)
        copy_body_style_from_left(ws, r, AK)
        ws.cell(row=r, column=AK).value = f"={rates_vlk(r)}"

        # 換算台幣 (AL) = M * AK
        copy_body_style_from_left(ws, r, AL)
        ws.cell(row=r, column=AL).value = f"={ws.cell(row=r, column=M).coordinate}*{ws.cell(row=r, column=AK).coordinate}"

        # 關係企業名稱 (AM) — resolve via mapping (no external formula)
        copy_body_style_from_left(ws, r, AM)
        key = normalize_id(ws.cell(row=r, column=AJ).value)
        ws.cell(row=r, column=AM).value = relparty_map.get(key, "")

    # accounting format for 換算台幣
    apply_accounting_format(ws, AL, 2, lr)




def append_calc_columns_43(ws, period: str, rates_path: Path, relparty_map: dict):
    from openpyxl.utils import column_index_from_string as colidx
    X, Y, Z = colidx("X"), colidx("Y"), colidx("Z")
    J, K, L = colidx("J"), colidx("K"), colidx("L")
    W = colidx("W")

    # headers
    copy_header_style(ws, W, X); ws.cell(row=1, column=X).value = "匯率"
    copy_header_style(ws, W, Y); ws.cell(row=1, column=Y).value = "換算台幣"
    copy_header_style(ws, W, Z); ws.cell(row=1, column=Z).value = "關係企業名稱"

    lr = last_data_row(ws, key_col=1)

    rates_vlk = lambda r: f"IF({ws.cell(row=r, column=K).coordinate}=\"NTD\",1," + \
        build_ext_vlookup(rates_path, "Summary", "$B:$C", ws.cell(row=r, column=K).coordinate, 2) + ")"

    for r in range(2, lr + 1):
        # 匯率 (X)
        copy_body_style_from_left(ws, r, X)
        ws.cell(row=r, column=X).value = f"={rates_vlk(r)}"

        # 換算台幣 (Y) = L * X
        copy_body_style_from_left(ws, r, Y)
        ws.cell(row=r, column=Y).value = f"={ws.cell(row=r, column=L).coordinate}*{ws.cell(row=r, column=X).coordinate}"

        # 關係企業名稱 (Z) — resolve via mapping
        copy_body_style_from_left(ws, r, Z)
        key = normalize_id(ws.cell(row=r, column=J).value)
        ws.cell(row=r, column=Z).value = relparty_map.get(key, "")

    # accounting format for 換算台幣
    apply_accounting_format(ws, Y, 2, lr)




# ---------- CLI ----------
def main():
    parser = argparse.ArgumentParser(description="Fill updated YTM forms by direct copy/paste with styles.")
    parser.add_argument("--template", required=True, help="Path to the template workbook (will be copied unless --inplace).")
    parser.add_argument("--out", help="Output path (.xlsx). If omitted, writes to default unless --inplace.")
    parser.add_argument("--inplace", action="store_true", help="Overwrite template in-place.")

    parser.add_argument(
        "--announce-sheet",
        default="1-1.公告(元)",   # change default if you prefer the dotted one
        help="Exact name of the sheet to update for 公告(月) structure."
    )
    parser.add_argument("--task", required=True,
                    choices=["copy_4_3", "copy_2_3", "both", "announce_structure", "all"],
                    help="Which sheet(s) to fill.")
    parser.add_argument("--src-43", help="Source workbook for 4-3 (columns B:X).")
    parser.add_argument("--src-23", help="Source workbook for 2-3 (columns A:AJ).")

    # NEW: external lookups
    parser.add_argument("--period", required=True, help="e.g., 202504")
    parser.add_argument("--rates-path", help="External rates workbook path; overrides default pattern.")
    parser.add_argument("--relparty-path", help="External related-party master workbook path.")
    parser.add_argument(
        "--rptis10-path",
        help="External RPTIS10 workbook path. Defaults to .../ytm_forms/data/template/關係人/RPTIS10_I_A01_<period>.xlsx"
    )

    parser.add_argument(
        "--mrs0034-path",
        help="External MRS0034 workbook path. Defaults to .../ytm_forms/data/template/關係人/MRS0034_I_A01_<period>.xlsx"
    )

    parser.add_argument(
        "--mrs0014-path",
        help="External MRS0014 workbook path. Defaults to .../ytm_forms/data/template/關係人/MRS0014_I_A01_<period>.xlsx"
    )

    args = parser.parse_args()
    # Base folder inside the repo
    BASE_TPL = PROJECT_ROOT / "ytm_forms" / "data" / "template" / "關係人"

    # Resolve sources (allow override via CLI)
    src_43 = Path(args.src_43) if args.src_43 else BASE_TPL / "export_關係人交易-應收帳款.xlsx"
    src_23 = Path(args.src_23) if args.src_23 else BASE_TPL / "export_關係人交易-收入.xlsx"

    # Resolve external files (allow override via CLI)
    yyyymm = args.period

    mrs0034_path = Path(args.mrs0034_path) if args.mrs0034_path else (
        BASE_TPL / f"MRS0034_I_A01_{yyyymm}.xlsx"
    )

    mrs0014_path = Path(args.mrs0014_path) if args.mrs0014_path else (
        BASE_TPL / f"MRS0014_I_A01_{yyyymm}.xlsx"
    )


    default_rates = BASE_TPL / f"{yyyymm} Ending 及 Avg (資通版本).xls"
    rates_path   = Path(args.rates_path) if args.rates_path else default_rates
    relparty_path = Path(args.relparty_path) if args.relparty_path else (BASE_TPL / "關係企業(人).xls")


    # validate template
    template_path = Path(args.template)
    if not template_path.exists():
        raise FileNotFoundError(f"Template not found: {template_path}")

    # resolve output path to ytm_forms/data/output (anchored to project root)
    out_path = Path(args.out) if args.out else (
        template_path if args.inplace else OUTPUT_DIR / f"copy_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    )
    final_out = load_output_from_template(template_path, out_path, args.inplace)

    if args.task in ("copy_4_3", "both", "all"):
        copy_43(template_path, src_43, final_out)
    if args.task in ("copy_2_3", "both", "all"):
        copy_23(template_path, src_23, final_out)
    
    # structure-only step for 1-1.公告(元) (no formulas yet)
    rptis10_path = Path(args.rptis10_path) if args.rptis10_path else (
        BASE_TPL / f"RPTIS10_I_A01_{args.period}.xlsx"
    )

    if args.task in ("announce_structure", "all"):
        wb_tmp = prepare_month_structure(final_out, sheet_name=args.announce_sheet, period_yyyymm=args.period, rptis10_path=rptis10_path, mrs0034_path=mrs0034_path, mrs0014_path=mrs0014_path)
        wb_tmp.save(final_out)
        wb_tmp.close()

    # warnings if links missing (optional)
    if not rates_path.exists():
        print(f"[WARN] Rates workbook not found: {rates_path}")
    if not relparty_path.exists():
        print(f"[WARN] Related-party workbook not found: {relparty_path}")

    # build {ID -> Name} map once
    relparty_map = load_relparty_map(relparty_path)


    # append 3 columns (匯率、換算台幣、關係企業名稱) with external formulas
    wb_final = load_workbook(final_out, data_only=False)

    if "4-3.應收關係人科餘" in wb_final.sheetnames:
        append_calc_columns_43(wb_final["4-3.應收關係人科餘"], args.period, rates_path, relparty_map)

    if "2-3.銷貨明細" in wb_final.sheetnames:
        append_calc_columns_23(wb_final["2-3.銷貨明細"], args.period, rates_path, relparty_map)


    wb_final.save(final_out)
    wb_final.close()

    print(f"Done → {final_out}")


if __name__ == "__main__":
    main()
