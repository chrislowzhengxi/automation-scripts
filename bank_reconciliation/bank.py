from collections import defaultdict
import shutil
from pathlib import Path
from parsers import CitiParser, CTBCParser, MegaParser, FubonParser, SinopacParser, ESunParser, BankParserBase
from fuzzy_matcher import match_entries_interactive, match_entries_debug
from utils import log_skipped

PARSER_REGISTRY = {
    "花旗": CitiParser,
    "中信": CTBCParser,
    "兆豐": MegaParser,
    "富邦": FubonParser,
    "永豐": SinopacParser,
    "玉山": ESunParser,
    # …more banks later…
}

def make_parser(path: Path) -> BankParserBase:
    stem = path.stem
    for key, cls in PARSER_REGISTRY.items():
        if key in stem:
            return cls(path)
    raise RuntimeError(f"No parser registered for {path.name}")


import argparse
import pandas as pd
import openpyxl
from datetime import datetime
from openpyxl.styles import Font
from rapidfuzz import process, fuzz
from pathlib import Path

# ─────────────── Configuration ───────────────
BASE_DIR        = Path("~/Downloads/Banks").expanduser()
BANK_FILE       = BASE_DIR / "花旗銀行對帳單-20250625.xlsx"
BANK_SHEET      = "Sheet2"
DB_FILE         = BASE_DIR / "會計憑證導入模板 - 1000 客戶資料庫.xls"
DB_SHEET        = "客戶資料庫"
FUZZY_THRESHOLD = 80
# OUTPUT_FILE = BASE_DIR / "會計憑證導入模板 - 空白檔案.xlsx"
RED_FONT    = Font(color="FF0000")


TEMPLATE_FILE = BASE_DIR / "會計憑證導入模板 - 空白檔案.xlsx"
def daily_output_path(post_date: str) -> Path:
    # post_date like "20250715"
    return BASE_DIR / f"會計憑證導入模板 - {post_date}.xlsx"

COL_DESC = "E"
COL_AMT  = "G"
KEYWORD  = "細節描述"

BANK_MAP = {
    "花旗": "花旗營業 NTD 0005",
    "中信": "中信營業 NTD 0800",
    "兆豐": "兆豐竹科新安 NTD 2656",
    "富邦": "富邦仁愛 NTD 6332",
    "永豐": "永豐城中 NTD 7978",
    "玉山": "玉山營業 NTD 8563",
    # …add more banks here…
}

# ─────────────── Functions ───────────────

def parse_args():
    p = argparse.ArgumentParser()
    p.add_argument(
        "--file", "-f",
        required=True,
        help="Path to the bank statement Excel file (xls or xlsx)"
    )
    p.add_argument("--date", "-d",
                   help="Posting date in YYYYMMDD (defaults to today)")
    
    p.add_argument(
        "--new-run",
        action="store_true",
        help="Start a new versioned file for this date instead of appending to the latest one."
    )
    return p.parse_args()

def detect_bank(stem, bank_map):
    for key, display in bank_map.items():
        if key in stem:
            print(f"Detected bank: '{display}' (matched '{key}')")
            return display
    raise RuntimeError(f"Cannot detect bank from filename: {stem!r}")

def load_and_filter_db(db_path, sheet, bank_display):
    # read .xls via pandas + xlrd
    df = pd.read_excel(db_path, sheet_name=sheet, engine="xlrd", header=None)
    df.columns = list("ABCDEFGHIJKLMNOPQRSTUVWXYZ")[:df.shape[1]]
    # filter on Column B containing the bank_display text
    filtered = df[df["B"].astype(str).str.contains(bank_display)]
    print(f"Filtered DB to {len(filtered)} rows for '{bank_display}'")
    return filtered


def day_output_base(post_date: str) -> Path:
    return BASE_DIR / f"會計憑證導入模板 - {post_date}.xlsx"

def enumerate_existing_outputs(post_date: str) -> list[Path]:
    """
    Returns existing files for the day in order:
    [base, -2, -3, ...] if they exist.
    """
    files = []
    base = day_output_base(post_date)
    if base.exists():
        files.append(base)
        k = 2
        while True:
            p = BASE_DIR / f"會計憑證導入模板 - {post_date}-{k}.xlsx"
            if p.exists():
                files.append(p)
                k += 1
            else:
                break
    return files


def latest_or_new_output_path(post_date: str, force_new_run: bool = False) -> tuple[Path, list[Path]]:
    """
    Default: append to the latest existing output for this date (base or -N).
    Only create a new -N file when force_new_run=True.
    Returns (out_path_to_write, earlier_paths_for_duplicate_check).
    """
    earlier = enumerate_existing_outputs(post_date)
    base = day_output_base(post_date)

    if not earlier:
        # No file yet for this date → create/use base
        return base, []

    if force_new_run:
        next_idx = len(earlier) + 1
        return BASE_DIR / f"會計憑證導入模板 - {post_date}-{next_idx}.xlsx", earlier

    # Reuse the latest existing file (append), whether it's base or a -N
    latest = earlier[-1]
    return latest, earlier


def collect_existing_counts(paths: list[Path]) -> dict:
    """
    Aggregate duplicate keys across ALL earlier files of the same day.
    Key = (E posting_date as str, U cust_id as str, S amount as float)
    """
    import openpyxl
    counts = defaultdict(int)
    for p in paths:
        wb = openpyxl.load_workbook(p, data_only=True)
        ws = wb["Sheet1"]
        for r in range(5, ws.max_row + 1, 2):
            e_val = ws.cell(r, 5).value   # E
            u_val = ws.cell(r, 21).value  # U
            s_val = ws.cell(r, 19).value  # S
            if e_val is None and u_val is None and s_val is None:
                continue
            try:
                s_float = float(str(s_val).replace(",", "")) if s_val is not None else 0.0
            except ValueError:
                s_float = 0.0
            key = (str(e_val), str(u_val), s_float)
            counts[key] += 1
    return counts


def write_output(matches, out_path: Path, post_date: str, existing_counts: dict) -> int:
    """
    Write ONLY new 2-row blocks into a NEW per-run workbook.
    We compare against aggregated counts from all earlier files (existing_counts).
    Returns the number of rows written (should be even).
    """
    import openpyxl

    # Create the per-run file from template
    if not out_path.exists():
        shutil.copy(TEMPLATE_FILE, out_path)

    wb = openpyxl.load_workbook(out_path)
    ws = wb["Sheet1"]

    def row_is_empty(r: int) -> bool:
        key_cols = [2,3,4,5,6,7,9,10,15,19,21,22]  # B,C,D,E,F,G,I,J,O,S,U,V
        return all(ws.cell(r, c).value is None for c in key_cols)

    # find first empty 2-row block starting at row 5
    row = 5
    while row <= ws.max_row + 1 and not row_is_empty(row):
        row += 2

    ymd    = post_date
    y_str  = ymd[:4]
    m_str  = ymd[4:6]
    md_str = f"{ymd[4:6]}.{ymd[6:]}"

    # def fill(r, data, red_cols=None):
    #     red_cols = red_cols or []
    #     for col, val in data.items():
    #         cell = ws[f"{col}{r}"]
    #         cell.value = val
    #         if col in red_cols:
    #             cell.font = RED_FONT
    #         if col == "S":
    #             cell.number_format = "#,##0.00"
    def fill(r, data, red_cols=None):
        red_cols = red_cols or []
        for col, val in data.items():
            cell = ws[f"{col}{r}"]
            # ensure amounts are real numbers
            if col == "S" and isinstance(val, str):
                try:
                    val = float(val.replace(",", ""))
                except Exception:
                    pass
            cell.value = val
            # if col == "S":
            #     # make sure Excel sees it as a number
            #     assert isinstance(cell.value, (int, float)), f"Cell {col}{r} is not numeric: {cell.value!r}"
            if col in red_cols:
                cell.font = RED_FONT
            if col == "S":
                # no thousands separator — SAP-friendly
                cell.number_format = "0.00"

    seen_now = defaultdict(int)
    written = 0

    for raw_txt, amt, db_row in matches:
        try:
            amt_float = float(str(amt).replace(",", "")) if amt is not None else 0.0
        except ValueError:
            amt_float = 0.0

        cust_id  = db_row["F"]
        clean_nm = db_row["G"]
        hkont    = db_row["C"]
        extra_H  = db_row["H"]
        extra_I  = db_row["I"]

        text_I = f"{md_str} {clean_nm} 暫收款"

        key = (ymd, str(cust_id), amt_float)
        seen_now[key] += 1

        # Skip until we exceed what’s already written in earlier files
        if seen_now[key] <= existing_counts.get(key, 0):
            continue

        # Row 1 (DZ)
        r1 = {
            "B": "1000", "C": y_str, "D": "DZ",
            "E": ymd,    "F": ymd,   "G": m_str,
            "I": text_I,
            "J": "NTD",  "O": hkont, "S": amt_float,
            "U": cust_id, "V": text_I,
            "AP": extra_H, "AU": extra_I,
        }
        fill(row, r1, red_cols=["E", "F", "G", "S"])

        # Row 2 (N=5)
        r2 = {
            "L": cust_id,
            "N": "5",
            "S": -amt_float,
            "U": cust_id,
            "V": text_I,
        }
        fill(row + 1, r2)

        row += 2
        written += 2

    wb.save(out_path)
    print(f"Wrote {written} rows into {out_path.name}")

    # Checking numeric properties 
    check_wb = openpyxl.load_workbook(out_path, data_only=True)
    check_ws = check_wb["Sheet1"]

    for row in range(5, check_ws.max_row + 1):
        val = check_ws[f"S{row}"].value
        if val is None:
            continue
        print(f"Row {row}, type={type(val)}, value={val}")

    return written

def ensure_xls_copy(xlsx_path: Path) -> Path:
    """
    Create an .xls copy of the given .xlsx using Excel automation.
    Returns the .xls path. If Excel/pywin32 is unavailable, raises ImportError.
    """
    from time import sleep
    try:
        import win32com.client as win32
    except ImportError as e:
        raise ImportError("pywin32 is required for .xls conversion. pip install pywin32") from e

    xls_path = xlsx_path.with_suffix(".xls")

    excel = win32.gencache.EnsureDispatch("Excel.Application")
    excel.Visible = False
    excel.DisplayAlerts = False
    try:
        wb = excel.Workbooks.Open(str(xlsx_path))
        # 56 = xlWorkbookNormal (Excel 97-2003, .xls)
        wb.SaveAs(str(xls_path), FileFormat=56)
        wb.Close(SaveChanges=False)
        # Give Excel a tick to flush
        sleep(0.1)
    finally:
        excel.Quit()

    return xls_path



def main():
    args      = parse_args()
    print(f"[ARGS] file={args.file} date={args.date or datetime.today().strftime('%Y%m%d')} new_run={args.new_run}")

    post_date = args.date or datetime.today().strftime("%Y%m%d")

    bank_path = Path(args.file).expanduser()
    parser    = make_parser(bank_path)
    entries   = parser.extract_rows()
    print(f"Loaded {len(entries)} entries from {bank_path.name}")

    # # 2) Detect which bank we’re processing
    # stem = Path(BANK_FILE).stem
    bank_display = detect_bank(bank_path.stem, BANK_MAP)

    # 3) Load & filter the customer DB
    db = load_and_filter_db(DB_FILE, DB_SHEET, bank_display)

    # # 4) Match
    # matches = match_entries_debug(entries, db, FUZZY_THRESHOLD)
    matches, skipped = match_entries_interactive(entries, db, FUZZY_THRESHOLD)
    if skipped:
        log_skipped(skipped, filepath="skipped.csv")

    print(f"DEBUG  → matches found: {len(matches)}")

    print("[INFO] Checking existing outputs & deciding target file...")
    out_path, earlier_paths = latest_or_new_output_path(post_date, force_new_run=args.new_run)
    print(f"[INFO] earlier_paths={ [p.name for p in earlier_paths] }")
    print(f"[MODE] {'NEW RUN' if args.new_run else 'Append to latest'}")
    print(f"[INFO] chosen out_path={out_path.name} (force_new_run={args.new_run})")


    preexisted = out_path.exists()  # track if we are appending to an existing -N

    existing_counts = collect_existing_counts(earlier_paths)

    # Write only new items to this file (append if it already exists)
    written = write_output(matches, out_path, post_date, existing_counts)

    # If nothing new was written, only delete if we created a brand new file this time
    if written == 0 and not preexisted:
        try:
            if args.new_run:
                # Keep the anchor file so subsequent files in this batch append to -N.
                print("No new entries; keeping the new per-run file as the batch anchor.")
            else:
                if out_path.exists():
                    out_path.unlink()
                print("No new entries; not creating a new per-run file.")
        except Exception as e:
            print(f"Note: could not remove empty file {out_path.name}: {e}")

    try:
        if out_path.exists():
            xls_out = ensure_xls_copy(out_path)
            print(f"[SAP] Also saved legacy Excel: {xls_out.name}")
            # remove the .xlsx after conversion
            try:
                out_path.unlink()
                print(f"[CLEANUP] Removed intermediate {out_path.name}")
            except Exception as e:
                print(f"[CLEANUP] Could not remove {out_path.name}: {e}")

    except ImportError as e:
        print(f"[SAP] Note: {e}. Skipping .xls creation.")
    except Exception as e:
        print(f"[SAP] Could not create .xls copy: {e}")

if __name__ == "__main__":
    main()
