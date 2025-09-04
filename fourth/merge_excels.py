#!/usr/bin/env python3
import argparse
import sys
from pathlib import Path
import pandas as pd

from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from unicodedata import east_asian_width


from group_by_gl import group_export_by_account, choose_output_path
from datetime import datetime, date


from pathlib import Path
import pandas as pd


def _text_display_units(s: str) -> float:
    """
    Approximate Excel display width in 'character' units.
    Count wide CJK as ~1.7x.
    """
    if s is None:
        return 0.0
    s = str(s)
    total = 0.0
    for ch in s:
        total += 1.7 if east_asian_width(ch) in ("W", "F") else 1.0
    return total

def enforce_arial_font(wb):
    """
    Switch only the font family to Arial, keep size/bold/italic/color/etc.
    """
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                f = cell.font
                # Only change if needed; Font is immutable → create a new one.
                if f is None or f.name != "Arial":
                    cell.font = Font(
                        name="Arial",
                        sz=f.sz if f else None,
                        b=f.b if f else None,
                        i=f.i if f else None,
                        color=f.color if f else None,
                        underline=f.u if f else None,
                        strike=f.strike if f else None,
                        vertAlign=f.vertAlign if f else None
                    )

def autofit_columns(ws, min_width=6, max_width=60, padding=2.0, skip_widths={1}):
    """
    Auto-size visible columns based on header + cell content.
    - Respects hidden columns.
    - Skips columns explicitly set to very narrow spacer widths (e.g., 1).
    - Preserves merged cells/number formats/wrap.
    """
    dims = ws.column_dimensions
    # Figure header row (first row with any value; fallback to row 1)
    header_row_idx = 1
    try:
        for r in range(1, ws.max_row + 1):
            if any(ws.cell(r, c).value not in (None, "") for c in range(1, ws.max_column + 1)):
                header_row_idx = r
                break
    except Exception:
        header_row_idx = 1

    for col_idx in range(1, ws.max_column + 1):
        col_letter = ws.cell(row=1, column=col_idx).column_letter
        cd = dims[col_letter] if col_letter in dims else None

        # Skip hidden columns
        if cd and getattr(cd, "hidden", False):
            continue
        # Skip spacer columns: width explicitly set to ~1
        if cd and cd.width is not None and round(cd.width, 1) in {float(w) for w in skip_widths}:
            continue

        # Measure header + all non-empty cells
        max_units = 0.0
        header_val = ws.cell(header_row_idx, col_idx).value
        max_units = max(max_units, _text_display_units(header_val))

        for row in ws.iter_rows(min_row=header_row_idx + 1,
                                max_row=ws.max_row,
                                min_col=col_idx, max_col=col_idx):
            cell = row[0]
            val = cell.value
            if val is None or val == "":
                continue
            # Use number/date displayed text length approximation
            text = str(val)
            max_units = max(max_units, _text_display_units(text))

        # Convert to Excel width units (rough heuristic)
        desired = max_units + padding
        desired = max(min_width, min(max_width, desired))

        # Don’t shrink columns that already have a larger explicit width
        if cd and cd.width and cd.width > desired:
            continue

        dims[col_letter].width = desired


def _first_numeric_format(ws, col_letter, start_row=2):
    for r in range(start_row, ws.max_row + 1):
        v = ws[f"{col_letter}{r}"].value
        if isinstance(v, (int, float)):
            fmt = ws[f"{col_letter}{r}"].number_format
            if fmt and fmt != "General":
                return fmt
    return None

def _apply_number_format(ws, col_letter, number_format, start_row=2):
    for r in range(start_row, ws.max_row + 1):
        c = ws[f"{col_letter}{r}"]
        if isinstance(c.value, (int, float)):
            c.number_format = number_format

def align_col_L_to_col_N(ws):
    fmt = _first_numeric_format(ws, "N") or "#,##0;[Red](#,##0)"
    _apply_number_format(ws, "L", fmt)


def _peek_bytes(path: Path, n=4096) -> bytes:
    with open(path, "rb") as f:
        return f.read(n)

def read_first_sheet(path: Path) -> pd.DataFrame:
    p = Path(path)
    head = _peek_bytes(p)
    suffix = p.suffix.lower()

    OLE2 = b"\xD0\xCF\x11\xE0"   # real .xls
    ZIP  = b"PK\x03\x04"         # real .xlsx
    html_markers = (b"<html", b"<!doctype", b"mhtml", b"mime-ver", b"content-type:", b"<table")
    xml_markers  = (b"<?xml", b"<Workbook", b"urn:schemas-microsoft-com:office:spreadsheet")

    # 1) True Excel
    if head.startswith(ZIP) or suffix == ".xlsx":
        return pd.read_excel(p, sheet_name=0, engine="openpyxl")
    if head.startswith(OLE2) and suffix == ".xls":
        return pd.read_excel(p, sheet_name=0, engine="xlrd")

    # 2) HTML/MHTML "fake .xls"
    if any(k in head.lower() for k in html_markers):
        tables = pd.read_html(p, header=0)  # take first row as header
        df = max(tables, key=lambda t: t.shape[1])  # widest table
        if df.shape[1] == 1:
            # Sometimes it's actually tab-delimited text in <pre>
            try:
                return pd.read_csv(p, sep="\t", engine="python", encoding="utf-8-sig")
            except Exception:
                try:
                    return pd.read_csv(p, sep="\t", engine="python", encoding="cp950")
                except Exception:
                    pass
        return df

    # 3) Excel 2003 XML (rare)
    if any(k in head for k in xml_markers):
        # Often better to open in Excel and Save As .xlsx; we can try xml:
        try:
            return pd.read_xml(p)
        except Exception as e:
            raise ValueError(f"Excel 2003 XML detected; please Save As .xlsx. ({e})")

    # 4) Try tab-delimited / CSV fallbacks
    for enc in ("utf-8-sig", "cp950"):
        try:
            return pd.read_csv(p, sep="\t", engine="python", encoding=enc)
        except Exception:
            pass
    for enc in ("utf-8-sig", "cp950"):
        try:
            return pd.read_csv(p, encoding=enc)
        except Exception:
            pass

    # 5) Last resort
    return pd.read_excel(p, sheet_name=0)


def normalize_columns(df: pd.DataFrame) -> pd.DataFrame:
    df.columns = [str(c).strip() for c in df.columns]
    return df

def merge_dataframes(dfs, ref_cols, drop_duplicates=True):
    # Keep only columns in ref and follow exact order
    cleaned = []
    for df in dfs:
        # Add any missing ref cols as empty
        for col in ref_cols:
            if col not in df.columns:
                df[col] = pd.NA
        cleaned.append(df[ref_cols])
    merged = pd.concat(cleaned, ignore_index=True)
    if drop_duplicates:
        merged = merged.drop_duplicates()
    return merged

def collect_input_files(args) -> list[Path]:
    files = []
    if args.inputs:
        files.extend([Path(p) for p in args.inputs])
    if args.dir:
        base = Path(args.dir)
        pattern = args.pattern or "*.xlsx"
        files.extend(sorted(base.glob(pattern)))
    # De-dup while preserving order
    seen = set()
    unique = []
    for f in files:
        if f not in seen:
            seen.add(f)
            unique.append(f)
    return unique

def main_cli():
    ap = argparse.ArgumentParser(description="Merge Excel files into a single workbook compatible with group_by_gl.py")
    ap.add_argument("--ref", required=False, help="Reference Excel file to enforce column order (recommended)")
    ap.add_argument("--inputs", nargs="*", help="Input Excel files")
    ap.add_argument("--dir", help="Directory to search for input Excel files")
    ap.add_argument("--pattern", help="Glob pattern within --dir, e.g. *.xlsx")
    ap.add_argument("--out", default="combined.xlsx", help="Output Excel file (default: combined.xlsx)")
    ap.add_argument("--sheet-name", default=None, help="Specific sheet name to read (default: first sheet)")
    ap.add_argument("--keep-duplicates", action="store_true", help="Keep duplicate rows (default: drop exact duplicates)")
    ap.add_argument("--gui", action="store_true", help="Open a simple GUI to pick files")
    ap.add_argument("--cutoff", default=None, help="Cutoff date (YYYY-MM-DD) for 30/90-day tests. Defaults to today.")
    args = ap.parse_args()


    # Parse cutoff date
    if args.cutoff:
        try:
            cutoff_date = datetime.strptime(args.cutoff, "%Y-%m-%d").date()
        except ValueError:
            print("[ERROR] --cutoff must be YYYY-MM-DD (e.g., 2025-06-30)", file=sys.stderr)
            sys.exit(2)
    else:
        cutoff_date = date.today()


    if args.gui:
        return run_gui()

    # Validate inputs
    inputs = collect_input_files(args)
    if not inputs:
        print("No input files specified. Use --inputs or --dir/--pattern, or run with --gui.", file=sys.stderr)
        sys.exit(2)

    # 🔍 Debug preview for each input
    for f in inputs:
        df = read_first_sheet(Path(f))
        df = normalize_columns(df)
        print(f"[DEBUG] {Path(f).name}: shape={df.shape}")
        print(df.head(2).to_string(index=False))

    # Load reference file (strongly recommended)
    if args.ref:
        ref_df = read_first_sheet(Path(args.ref))
    else:
        # If no reference, use the first input as reference
        ref_df = read_first_sheet(inputs[0])
        print(f"[Info] No --ref provided. Using first input as reference: {inputs[0].name}")

    ref_df = normalize_columns(ref_df)
    ref_cols = list(ref_df.columns)

    # Load all input files
    dataframes = []
    for f in inputs:
        df = read_first_sheet(f)
        df = normalize_columns(df)
        dataframes.append(df)

    merged = merge_dataframes(dataframes, ref_cols, drop_duplicates=not args.keep_duplicates)

    # 👉 Format date columns
    for col in ["文件日期", "過帳日期"]:
        if col in merged.columns:
            merged[col] = pd.to_datetime(merged[col], errors="coerce").dt.strftime("%Y-%m-%d")

    # Write output (single sheet)
    out_path = Path(args.out)
    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        merged.to_excel(writer, index=False, sheet_name="Sheet1")

    print(f"✅ Merged {len(inputs)} files → {out_path}")

    # === Immediately run grouping ===
    mapping_path = Path("會計科目對照表.xlsx")  # adjust path if needed
    output_path = out_path.with_name(out_path.stem + "_grouped.xlsx")

    stats = group_export_by_account(
        export_path=out_path,
        mapping_path=mapping_path,
        output_path=output_path,
        sheet_name=None,
        inplace=False,
        drop_original_titles=["Sheet1","Sheet2","Sheet3"],
        date_columns=["文件日期","過帳日期"],
        cutoff_date=cutoff_date
    )

    print(f"✅ Grouped output written to {output_path}")

    # --- Post-merge formatting: Arial + Autofit ---
    # Format the merged file
    wb = load_workbook(out_path)
    enforce_arial_font(wb)
    for ws in wb.worksheets:
        autofit_columns(ws, min_width=6, max_width=60, padding=2.0, skip_widths={1})
    wb.save(out_path)

    # Format the grouped file
    gwb = load_workbook(output_path)
    enforce_arial_font(gwb)
    for ws in gwb.worksheets:
        if ws.title.startswith("說明"):   # skip documentation sheets
            continue
        align_col_L_to_col_N(ws)          # <-- add this
        autofit_columns(ws, min_width=6, max_width=60, padding=2.0, skip_widths={1})
    gwb.save(output_path)

    for k, v in stats.items():
        print(f"- {k}: {v}")



def run_gui():
    import tkinter as tk
    from tkinter import filedialog, messagebox

    root = tk.Tk()
    root.title("Merge Excel files for group_by_gl.py")
    root.geometry("560x320")

    state = {"ref": None, "inputs": []}

    def choose_ref():
        p = filedialog.askopenfilename(
            title="Choose reference Excel (enforces column order)",
            filetypes=[("Excel files", "*.xlsx *.xls")]
        )
        if p:
            state["ref"] = p
            ref_var.set(p)

    def choose_inputs():
        ps = filedialog.askopenfilenames(
            title="Choose input Excel files to merge",
            filetypes=[("Excel files", "*.xlsx *.xls")]
        )
        if ps:
            state["inputs"] = list(ps)
            inputs_var.set("\n".join(ps))

    def do_merge():
        try:
            if not state["inputs"]:
                messagebox.showerror("Error", "Please choose at least one input file.")
                return
            out = out_entry.get().strip() or "combined.xlsx"

            cutoff_str = cutoff_entry.get().strip()
            if cutoff_str:
                try:
                    cutoff_dt = datetime.strptime(cutoff_str, "%Y-%m-%d").date()
                except ValueError:
                    messagebox.showerror("Error", "Cutoff date must be YYYY-MM-DD (e.g., 2025-06-30).")
                    return
            else:
                cutoff_dt = date.today()        

            # Load reference (or fallback)
            if state["ref"]:
                ref_df = read_first_sheet(Path(state["ref"]))
            else:
                ref_df = read_first_sheet(Path(state["inputs"][0]))

            ref_df = normalize_columns(ref_df)
            ref_cols = list(ref_df.columns)

            # 👉 Auto-include the reference file in inputs if not already selected
            files = list(state["inputs"])
            if state["ref"] and state["ref"] not in files:
                files.insert(0, state["ref"])

            dfs = []
            for p in files:
                df = read_first_sheet(Path(p))
                df = normalize_columns(df)
                dfs.append(df)

            merged = merge_dataframes(dfs, ref_cols, drop_duplicates=(not keep_dups_var.get()))
            
            # 👈 paste here
            for col in ["文件日期", "過帳日期"]:
                if col in merged.columns:
                    merged[col] = pd.to_datetime(merged[col], errors="coerce").dt.strftime("%Y-%m-%d")

            out_path = Path(out)
            with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
                merged.to_excel(writer, index=False, sheet_name="Sheet1")


            # === Immediately run group_by_gl.py on merged file ===
            from group_by_gl import group_export_by_account


            mapping_path = Path("會計科目對照表.xlsx")  # adjust path if needed
            output_path = out_path.with_name(out_path.stem + "_grouped.xlsx")

            stats = group_export_by_account(
                export_path=out_path,
                mapping_path=mapping_path,
                output_path=output_path,
                sheet_name=None,
                inplace=False,
                drop_original_titles=["Sheet1", "Sheet2", "Sheet3"],
                date_columns=["文件日期", "過帳日期"],
                cutoff_date=cutoff_dt
            )


            # --- Post-merge formatting: Arial + Autofit ---
            wb = load_workbook(out_path)
            enforce_arial_font(wb)
            for ws in wb.worksheets:
                autofit_columns(ws, min_width=6, max_width=60, padding=2.0, skip_widths={1})
            wb.save(out_path)

            gwb = load_workbook(output_path)
            enforce_arial_font(gwb)
            for ws in gwb.worksheets:
                if ws.title.startswith("說明"):
                    continue
                align_col_L_to_col_N(ws)  
                autofit_columns(ws, min_width=6, max_width=60, padding=2.0, skip_widths={1})
            gwb.save(output_path)

            # Nice feedback: show how many went in and how many rows came out
            import tkinter
            messagebox.showinfo(
                "Done",
                f"Merged {len(files)} files → {out_path}\n"
                f"Rows: {len(merged):,}   Columns: {len(merged.columns)}"
            )
        except Exception as e:
            messagebox.showerror("Error", str(e))


    # UI
    frm = tk.Frame(root, padx=10, pady=10)
    frm.pack(fill="both", expand=True)

    ref_var = tk.StringVar()
    inputs_var = tk.StringVar()
    keep_dups_var = tk.BooleanVar(value=False)

    tk.Label(frm, text="Reference file (recommended):").grid(row=0, column=0, sticky="w")
    tk.Entry(frm, textvariable=ref_var, width=60).grid(row=1, column=0, columnspan=2, sticky="we")
    tk.Button(frm, text="Choose reference...", command=choose_ref).grid(row=1, column=2, padx=6)

    tk.Label(frm, text="Input Excel files to merge:").grid(row=2, column=0, sticky="w", pady=(10,0))
    tk.Entry(frm, textvariable=inputs_var, width=60).grid(row=3, column=0, columnspan=2, sticky="we")
    tk.Button(frm, text="Choose files...", command=choose_inputs).grid(row=3, column=2, padx=6)

    tk.Label(frm, text="Output file name:").grid(row=4, column=0, sticky="w", pady=(10,0))
    out_entry = tk.Entry(frm, width=30)
    out_entry.insert(0, "combined.xlsx")
    out_entry.grid(row=5, column=0, sticky="w")


    tk.Label(frm, text="Cutoff date (YYYY-MM-DD, blank = today):").grid(row=6, column=0, sticky="w", pady=(10,0))
    cutoff_entry = tk.Entry(frm, width=30)
    cutoff_entry.insert(0, "")  # leave blank => today
    cutoff_entry.grid(row=7, column=0, sticky="w")


    tk.Checkbutton(frm, text="Keep duplicate rows", variable=keep_dups_var).grid(row=8, column=0, sticky="w", pady=(10,0))

    tk.Button(frm, text="Merge", command=do_merge).grid(row=9, column=0, pady=12, sticky="w")
    tk.Button(frm, text="Close", command=root.destroy).grid(row=9, column=1, pady=12, sticky="w")

    root.mainloop()


if __name__ == "__main__":
    main_cli()
