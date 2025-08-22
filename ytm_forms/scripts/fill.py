#!/usr/bin/env python3
import argparse, os, re, shutil
from datetime import datetime, date 
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
from copy import copy

src = Path(r"C:\Users\TP2507088\Downloads\202504YTM WITS-C 上櫃公司與關係人間重要交易資訊.xlsx")
dst = src.with_name(f"Copy {src.name}")

shutil.copy(src, dst)
print(f"Copied to: {dst}")

TARGET_SHEET = "2-2.銷貨倍力"
COLUMN_A = 1
COLUMN_S = 19
PERIOD_RX = re.compile(r"(20\d{2})(0[1-9]|1[0-2])")

# ---------- common helpers ----------
def find_downloads() -> Path:
    up = os.environ.get("USERPROFILE")
    d = Path(up) / "Downloads" if up else Path.home() / "Downloads"
    return d

def pick_file_by_period(prefix: str, period: str, explicit: str | None) -> Path:
    if explicit:
        p = Path(explicit)
        if not p.exists():
            raise FileNotFoundError(f"Path not found: {p}")
        return p

    downloads = find_downloads()
    candidates = sorted(Path(downloads).glob(f"*{prefix}*.xlsx"))
    if not candidates:
        raise FileNotFoundError(f"No {prefix} files in {downloads}")

    # 1) exact period match
    period_matches = [p for p in candidates if period in p.name]
    if period_matches:
        return max(period_matches, key=lambda p: p.stat().st_mtime)

    # 2) else pick highest YYYYMM we can parse
    def pnum(p: Path):
        m = PERIOD_RX.search(p.name)
        return int(m.group(1) + m.group(2)) if m else None

    with_period = [(p, pnum(p)) for p in candidates]
    with_period = [(p, n) for (p, n) in with_period if n is not None]
    if with_period:
        max_period = max(n for _, n in with_period)
        newest_in_max = max([p for (p, n) in with_period if n == max_period],
                            key=lambda p: p.stat().st_mtime)
        return newest_in_max

    # 3) fallback newest by mtime
    return max(candidates, key=lambda p: p.stat().st_mtime)

def copy_cell(src, dst):
    # value
    dst.value = src.value
    # number format + basic style so red/parentheses/accounting look right
    dst.number_format = src.number_format
    dst.font = copy(src.font)
    dst.fill = copy(src.fill)
    dst.border = copy(src.border)
    dst.alignment = copy(src.alignment)
    dst.protection = copy(src.protection)

def dest_anchor(ws, r, c):
    for rng in ws.merged_cells.ranges:
        if rng.min_row <= r <= rng.max_row and rng.min_col <= c <= rng.max_col:
            return rng.min_row, rng.min_col
    return r, c

def default_export_path(explicit: str | None) -> Path:
    if explicit:
        p = Path(explicit)
        if not p.exists():
            raise FileNotFoundError(f"export.xlsx not found: {p}")
        return p
    d = find_downloads()
    p = d / "export.xlsx"
    if not p.exists():
        raise FileNotFoundError(f"export.xlsx not found at {p}")
    return p


def quote_sheet(name: str) -> str:
    # always quote sheet names for safety
    return f"'{name}'" if not (name.startswith("'") and name.endswith("'")) else name

def build_ext_vlookup(path: Path, sheet: str, table_range: str, key_ref: str, col_index: int) -> str:
    # 'C:\...\[file.xls]Sheet'!$B:$C  → VLOOKUP(key, that, col_index, FALSE)
    book = f"[{path.name}]"
    sheet_quoted = quote_sheet(sheet)
    xref = f"'{str(path.parent)}\\{book}{sheet_quoted[1:-1]}'!{table_range}"
    return f"VLOOKUP({key_ref},{xref},{col_index},FALSE)"

def copy_header_style(ws, src_col_idx: int, dst_col_idx: int):
    s = ws.cell(row=1, column=src_col_idx)
    d = ws.cell(row=1, column=dst_col_idx)
    d.value = s.value  # temp; caller will overwrite with new header text
    d.number_format = s.number_format
    from copy import copy as _cpy
    d.font = _cpy(s.font); d.fill = _cpy(s.fill); d.border = _cpy(s.border)
    d.alignment = _cpy(s.alignment); d.protection = _cpy(s.protection)
    # copy column width
    try:
        from openpyxl.utils import get_column_letter
        dst_letter = get_column_letter(dst_col_idx)
        src_letter = get_column_letter(src_col_idx)
        ws.column_dimensions[dst_letter].width = ws.column_dimensions[src_letter].width
    except Exception:
        pass

def copy_body_style_from_left(ws, row: int, col_idx: int):
    # style like the immediate left neighbor (common pattern in your sheets)
    if col_idx <= 1: 
        return
    s = ws.cell(row=row, column=col_idx - 1)
    d = ws.cell(row=row, column=col_idx)
    from copy import copy as _cpy
    d.number_format = s.number_format
    d.font = _cpy(s.font); d.fill = _cpy(s.fill); d.border = _cpy(s.border)
    d.alignment = _cpy(s.alignment); d.protection = _cpy(s.protection)


# ---------- task: MRS0014 → 2-2 ----------
def run_mrs0014(template_path: Path, period: str, mrs_path: str | None, out_path: Path):
    mrs_file = pick_file_by_period("MRS0014", period, mrs_path)

    # read values for 421007 / 421807 from column S
    wb_src = load_workbook(mrs_file, data_only=True)
    vals = {"421007": 0, "421807": 0}
    try:
        for ws in wb_src.worksheets:
            for r in range(1, (ws.max_row or 5000) + 1):
                a_val = ws.cell(row=r, column=COLUMN_A).value
                if a_val is None:
                    continue
                key = str(a_val).strip()
                if key in vals:
                    s_val = ws.cell(row=r, column=COLUMN_S).value or 0
                    try:
                        vals[key] += s_val
                    except TypeError:
                        pass
    finally:
        wb_src.close()

    # write to template (C18, C19, sum to C21)
    wb = load_workbook(template_path)
    try:
        ws = wb[TARGET_SHEET] if TARGET_SHEET in wb.sheetnames else wb.active
        ws["C18"] = vals["421007"]
        ws["C19"] = vals["421807"]
        ws["C21"] = "=SUM(C18:C19)"
        out_path.parent.mkdir(parents=True, exist_ok=True)
        wb.save(out_path)
    finally:
        wb.close()

# ---------- task: RPTIS10 → 2-2 ----------
def run_rptis10(template_path: Path, period: str, rpt_path: str | None, out_path: Path,
                src_sheet: str | None = None, src_rows=(6,11), src_cols=(1,3),
                dst_start_row=7, dst_start_col=1):
    rpt_file = pick_file_by_period("RPTIS10", period, rpt_path)

    wb_src = load_workbook(rpt_file, data_only=False)  # keep styles
    try:
        ws_src = wb_src[src_sheet] if (src_sheet and src_sheet in wb_src.sheetnames) else wb_src.active
        r1, r2 = src_rows
        c1, c2 = src_cols
        block = []
        for r in range(r1, r2+1):
            row_cells = []
            for c in range(c1, c2+1):
                row_cells.append(ws_src.cell(row=r, column=c))
            block.append(row_cells)
    finally:
        wb_src.close()

    wb = load_workbook(template_path)
    try:
        ws_dst = wb[TARGET_SHEET] if TARGET_SHEET in wb.sheetnames else wb.active
        # for i, row_cells in enumerate(block):
        #     dst_row = dst_start_row + i
        #     for j, src_cell in enumerate(row_cells):
        #         dst_col = dst_start_col + j
        #         copy_cell(src_cell, ws_dst.cell(row=dst_row, column=dst_col))
        anchors_written = set()
        for i, row_cells in enumerate(block):
            dst_row = dst_start_row + i
            for j, src_cell in enumerate(row_cells):
                dst_col = dst_start_col + j
                ar, ac = dest_anchor(ws_dst, dst_row, dst_col)
                key = (ar, ac)
                if key in anchors_written:
                    continue
                anchors_written.add(key)
                dst_cell = ws_dst.cell(row=ar, column=ac)
                copy_cell(src_cell, dst_cell)
        out_path.parent.mkdir(parents=True, exist_ok=True)
        wb.save(out_path)
    finally:
        wb.close()


# ------- 4.3 -----------
def _read_export_rows(ws_src, skip_header=True, drop_trailing=3):
    """
    Read rows from export.xlsx columns B..W (2..23).
    - skip_header: drops the first row
    - drop_trailing: drops this many trailing rows (e.g., totals)
    """
    rows = []
    empty_streak = 0
    started = False
    max_r = ws_src.max_row or 5000

    # start at row 2 if skipping header, else row 1
    start_row = 2 if skip_header else 1

    for r in range(start_row, max_r + 1):
        vals = [ws_src.cell(row=r, column=c).value for c in range(2, 24)]  # B..W
        if any(v is not None and str(v).strip() != "" for v in vals):
            rows.append(vals)
            started = True
            empty_streak = 0
        else:
            if started:
                empty_streak += 1
                if empty_streak >= 100:
                    break

    # drop trailing total rows
    if drop_trailing and len(rows) > drop_trailing:
        rows = rows[:-drop_trailing]

    return rows



def run_export_paste(template_path: Path, export_path: Path, out_path: Path,
                     dest_sheet: str = "4-3.應收關係人科餘", dst_start_row: int = 10):
    """
    Copy export.xlsx B..W → template A..V on the 4-3 sheet, then:
      - set W=1,
      - set X = N*W (as a formula),
      - format columns E and F as YYYY/MM/DD (strip time).
    Inserts new rows at dst_start_row so nothing above is overwritten.
    """
    wb_dst = load_workbook(template_path)
    try:
        ws_dst = wb_dst[dest_sheet] if dest_sheet in wb_dst.sheetnames else wb_dst.active

        insert_at = dst_start_row  # lock to row 10 as discussed

        # --- load export data (skip header, drop last 3 totals) ---
        wb_src = load_workbook(export_path, data_only=True)
        try:
            ws_src = wb_src.active
            rows = _read_export_rows(ws_src, skip_header=True, drop_trailing=3)
        finally:
            wb_src.close()

        if not rows:
            raise RuntimeError("No data found in export.xlsx (after skipping header and totals).")

        # --- insert rows so we never overwrite existing content ---
        ws_dst.insert_rows(insert_at, amount=len(rows))

        # --- paste values and set W/X ---
        for i, vals in enumerate(rows):
            r = insert_at + i

            # export B..W -> dest A..V
            for j, v in enumerate(vals[:22]):
                ws_dst.cell(row=r, column=1 + j).value = v

            # W (23) = 1
            ws_dst.cell(row=r, column=23).value = 1
            # X (24) = N * W (formula)
            ws_dst.cell(row=r, column=24).value = f"=N{r}*W{r}"

            # ---- format E & F as dates (YYYY/MM/DD) and strip time ----
            for col in (5, 6):  # E, F
                c = ws_dst.cell(row=r, column=col)
                # if it's a datetime, replace with just the date
                if isinstance(c.value, datetime):
                    c.value = c.value.date()
                # if it's a string like '2025-08-08 00:00:00', best-effort parse
                elif isinstance(c.value, str):
                    try:
                        c.value = datetime.fromisoformat(c.value).date()
                    except Exception:
                        pass
                c.number_format = "yyyy/mm/dd"

        out_path.parent.mkdir(parents=True, exist_ok=True)
        wb_dst.save(out_path)
    finally:
        wb_dst.close()



def main():
    ap = argparse.ArgumentParser(description="Fill YTM forms")
    ap.add_argument("--task", required=True,
                    choices=["mrs0014", "rptis10", "both", "export_4_3"])
    ap.add_argument("--period", required=True, help="e.g., 202504")
    ap.add_argument("--template", required=True, help="Path to the template workbook to fill")

    mx = ap.add_mutually_exclusive_group()
    mx.add_argument("--out", help="Output path (.xlsx). If omitted, writes to default unless --inplace is set")
    mx.add_argument("--inplace", action="store_true", help="Overwrite the template file")

    ap.add_argument("--dest-start-row", type=int, default=10,
                    help="First row to paste into on destination sheet (export_4_3)")

    # 2-2 options
    ap.add_argument("--mrs", help="Explicit path to MRS0014 (optional)")
    ap.add_argument("--rptis", help="Explicit path to RPTIS10 (optional)")
    ap.add_argument("--rpt-source-sheet", help="RPTIS10 source sheet name (optional)")

    # 4-3 options
    ap.add_argument("--export", help="Path to export.xlsx (optional; defaults to Downloads/export.xlsx)")
    ap.add_argument("--dest-sheet", default="4-3.應收關係人科餘",
                    help="Destination sheet name for the export paste")

    args = ap.parse_args()

    template_path = Path(args.template)
    if not template_path.exists():
        raise FileNotFoundError(f"Template not found: {template_path}")

    out_path = Path(args.out) if args.out else (
        template_path if args.inplace else Path(f"ytm_forms/data/output/{args.task}_{args.period}.xlsx")
    )

    if args.task == "mrs0014":
        run_mrs0014(template_path, args.period, args.mrs, out_path)
    elif args.task == "rptis10":
        run_rptis10(template_path, args.period, args.rptis, out_path, src_sheet=args.rpt_source_sheet)
    elif args.task == "both":
        run_mrs0014(template_path, args.period, args.mrs, out_path)
        run_rptis10(template_path, args.period, args.rptis, out_path, src_sheet=args.rpt_source_sheet)
    else:  # export_4_3
        exp_path = default_export_path(args.export)
        run_export_paste(template_path, exp_path, out_path,
                         dest_sheet=args.dest_sheet, dst_start_row=args.dest_start_row)

    print(f"Done → {out_path}")


if __name__ == "__main__":
    main()

